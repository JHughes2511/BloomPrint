"""Excel upload and import for player/team evaluations."""

import io
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..auth import get_current_coach
from ..database import get_db
from .. import models

router = APIRouter(prefix="/uploads", tags=["uploads"])

# ── Header alias table ─────────────────────────────────────────────────────────
# Maps normalised column header → internal field name (pillar: prefix = pillar grade)

HEADER_MAP: dict[str, str] = {
    # Player identity
    "player name": "player_name", "name": "player_name",
    "player": "player_name", "athlete": "player_name",
    # Meta
    "position": "position", "pos": "position",
    "competition level": "competition_level", "level": "competition_level",
    "class": "competition_level",
    "output type": "output_type", "report type": "output_type", "type": "output_type",
    # Grades
    "overall grade": "overall_grade", "overall": "overall_grade",
    "grade": "overall_grade", "score": "overall_grade",
    "rating": "overall_grade", "total": "overall_grade",
    # Flags / questions
    "green flags": "green_flags", "strengths": "green_flags", "positives": "green_flags",
    "watch flags": "watch_flags", "concerns": "watch_flags", "weaknesses": "watch_flags",
    "red flags": "watch_flags", "negatives": "watch_flags",
    "key questions": "key_questions", "questions": "key_questions",
    # Report text
    "notes": "report_text", "report": "report_text", "comments": "report_text",
    "report text": "report_text", "scouting notes": "report_text",
    "evaluation": "report_text", "summary": "report_text",
    # Box score stats
    "pts": "stat:pts", "points": "stat:pts", "ppg": "stat:pts",
    "reb": "stat:reb", "rebounds": "stat:reb", "rpg": "stat:reb", "total reb": "stat:reb",
    "ast": "stat:ast", "assists": "stat:ast", "apg": "stat:ast",
    "stl": "stat:stl", "steals": "stat:stl", "spg": "stat:stl",
    "blk": "stat:blk", "blocks": "stat:blk", "bpg": "stat:blk",
    "to": "stat:to", "turnovers": "stat:to", "tov": "stat:to",
    "fgm": "stat:fgm", "fga": "stat:fga",
    "fg%": "stat:fg_pct", "fg pct": "stat:fg_pct", "field goal pct": "stat:fg_pct",
    "3pm": "stat:three_pm", "3pa": "stat:three_pa",
    "3p%": "stat:three_pct", "3pt pct": "stat:three_pct", "three pct": "stat:three_pct",
    "ftm": "stat:ftm", "fta": "stat:fta",
    "ft%": "stat:ft_pct", "free throw pct": "stat:ft_pct",
    "min": "stat:min", "minutes": "stat:min", "mpg": "stat:min",
    "gp": "stat:gp", "games played": "stat:gp", "games": "stat:gp",
    "+/-": "stat:plus_minus", "plus minus": "stat:plus_minus",
    "oreb": "stat:oreb", "off reb": "stat:oreb",
    "dreb": "stat:dreb", "def reb": "stat:dreb",
    # Pillars
    "offensive skills": "pillar:offensive_skills", "offensive": "pillar:offensive_skills",
    "offense": "pillar:offensive_skills", "off skills": "pillar:offensive_skills",
    "defensive capabilities": "pillar:defensive_capabilities",
    "defensive": "pillar:defensive_capabilities", "defense": "pillar:defensive_capabilities",
    "def skills": "pillar:defensive_capabilities",
    "physical attributes": "pillar:physical_attributes",
    "physical": "pillar:physical_attributes", "athleticism": "pillar:physical_attributes",
    "phys": "pillar:physical_attributes",
    "intangibles": "pillar:intangibles", "intangible": "pillar:intangibles",
    "iq": "pillar:intangibles", "character": "pillar:intangibles",
    "leadership": "pillar:intangibles",
    "advanced analysis": "pillar:advanced_analysis",
    "advanced": "pillar:advanced_analysis", "analytics": "pillar:advanced_analysis",
    "efficiency": "pillar:advanced_analysis",
    "strategic fit": "pillar:strategic_fit", "strategic": "pillar:strategic_fit",
    "fit": "pillar:strategic_fit", "system fit": "pillar:strategic_fit",
}


class ImportResult(BaseModel):
    players_created: int = 0
    players_found: int = 0
    evaluations_created: int = 0
    rows_processed: int = 0
    errors: list[str] = []


# ── Helpers ────────────────────────────────────────────────────────────────────

def _norm(h: Any) -> str:
    return str(h).lower().strip().replace("_", " ").replace("-", " ")


def _to_float(val: Any) -> float | None:
    try:
        return round(float(val), 2)
    except (TypeError, ValueError):
        return None


def _to_list(val: Any) -> list[str]:
    if val is None or str(val).strip() == "":
        return []
    return [x.strip() for x in str(val).replace(";", ",").split(",") if x.strip()]


def _parse_sheet(ws) -> list[dict]:
    """Return a list of row dicts from a worksheet using HEADER_MAP."""
    rows = list(ws.iter_rows(values_only=True))

    # Find first non-empty row → treat as header
    header_idx = next(
        (i for i, r in enumerate(rows) if any(c is not None and str(c).strip() for c in r)),
        None,
    )
    if header_idx is None:
        return []

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(rows[header_idx]):
        if cell is None:
            continue
        mapped = HEADER_MAP.get(_norm(cell))
        if mapped:
            col_map[idx] = mapped

    if not col_map:
        return []

    records: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if not any(c is not None and str(c).strip() for c in row):
            continue  # empty row

        record: dict[str, Any] = {}
        pillars: dict[str, float] = {}
        stats: dict[str, Any] = {}

        for idx, field in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is None or str(val).strip() == "":
                continue

            if field.startswith("pillar:"):
                f = _to_float(val)
                if f is not None:
                    pillars[field[7:]] = f
            elif field.startswith("stat:"):
                stats[field[5:]] = str(val).strip()
            elif field == "overall_grade":
                record[field] = _to_float(val)
            elif field in ("green_flags", "watch_flags", "key_questions"):
                record[field] = _to_list(val)
            else:
                record[field] = str(val).strip()

        if pillars:
            record["pillar_grades"] = pillars
        # Format box score stats into report_text if no text column provided
        if stats and not record.get("report_text"):
            stat_label = {
                "pts": "PTS", "reb": "REB", "ast": "AST", "stl": "STL", "blk": "BLK",
                "to": "TO", "fgm": "FGM", "fga": "FGA", "fg_pct": "FG%",
                "three_pm": "3PM", "three_pa": "3PA", "three_pct": "3P%",
                "ftm": "FTM", "fta": "FTA", "ft_pct": "FT%",
                "min": "MIN", "gp": "GP", "plus_minus": "+/-",
                "oreb": "OREB", "dreb": "DREB",
            }
            lines = "\n".join(
                f"- {stat_label.get(k, k.upper())}: {v}" for k, v in stats.items()
            )
            record["report_text"] = f"BOX SCORE STATS\n{lines}"
            if "output_type" not in record:
                record["output_type"] = "box_score"
        if record.get("player_name"):
            records.append(record)

    return records


# ── Endpoint ───────────────────────────────────────────────────────────────────

@router.post("/excel", response_model=ImportResult)
async def import_excel(
    file: UploadFile = File(...),
    output_type: str = Form("player_eval"),
    competition_level: str = Form("HS Varsity"),
    team_id: int | None = Form(None),
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    result = ImportResult()

    for sheet_name in wb.sheetnames:
        rows = _parse_sheet(wb[sheet_name])

        for row in rows:
            result.rows_processed += 1
            player_name = (row.get("player_name") or "").strip()
            if not player_name:
                continue

            try:
                # Find or create player
                player = (
                    db.query(models.Player)
                    .filter(models.Player.name.ilike(player_name))
                    .first()
                )

                if player is None:
                    program_name = coach.program_name
                    if team_id:
                        t = db.get(models.Team, team_id)
                        if t:
                            program_name = t.name
                    player = models.Player(
                        name=player_name,
                        position=row.get("position"),
                        competition_level=row.get("competition_level", competition_level),
                        program_name=program_name,
                        team_id=team_id,
                    )
                    db.add(player)
                    db.flush()
                    result.players_created += 1
                else:
                    result.players_found += 1
                    if not player.position and row.get("position"):
                        player.position = row["position"]
                    if row.get("competition_level"):
                        player.competition_level = row["competition_level"]

                # Auto-build minimal report_text from flags if no notes column
                report_text = row.get("report_text")
                if not report_text:
                    parts: list[str] = []
                    if row.get("green_flags"):
                        parts.append("GREEN FLAGS\n" + "\n".join(f"- {f}" for f in row["green_flags"]))
                    if row.get("watch_flags"):
                        parts.append("WATCH FLAGS\n" + "\n".join(f"- {f}" for f in row["watch_flags"]))
                    if row.get("key_questions"):
                        parts.append("KEY QUESTIONS\n" + "\n".join(f"- {q}" for q in row["key_questions"]))
                    report_text = "\n\n".join(parts) or None

                ev = models.Evaluation(
                    player_id=player.id,
                    coach_id=coach.id,
                    output_type=row.get("output_type", output_type),
                    competition_level=row.get("competition_level", competition_level),
                    coach_weight=coach.weight,
                    overall_grade=row.get("overall_grade"),
                    pillar_grades=row.get("pillar_grades"),
                    green_flags=row.get("green_flags"),
                    watch_flags=row.get("watch_flags"),
                    key_questions=row.get("key_questions"),
                    report_text=report_text,
                )
                db.add(ev)
                result.evaluations_created += 1

            except Exception as exc:
                result.errors.append(f"'{player_name}': {exc}")

    db.commit()
    return result

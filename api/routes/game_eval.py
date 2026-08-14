"""Game Evaluation routes — BIM game-by-game grading system."""

import difflib
import io
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File
from sqlalchemy import func
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..database import get_db, SessionLocal
from ..auth import get_current_coach
from ..report_format import REPORT_FORMAT, REPORT_FORMAT_WITH_TABLES
from .. import models, schemas
from ..softdelete import soft_delete
from ..ai_models import OPUS, SONNET, long_text
from ..uploadguard import read_upload
from .. import ai_import
from .. import genjob

router = APIRouter(prefix="/game-eval", tags=["game-eval"])

# ── Stat definitions ──────────────────────────────────────────────────────────

OFFENSE_STATS: dict[str, dict] = {
    "2 FG Made":      {"base_low": 2,  "base_high": 3,  "threshold": 4},
    "2 FG Missed":    {"base_low": -1, "base_high": -2, "threshold": 4},
    "3 FG Made":      {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "3 FG Missed":    {"base_low": -1, "base_high": -2, "threshold": 4},
    "Off. Reb":       {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "Draw PF":        {"base_low": 1,  "base_high": 1,  "threshold": 4},
    "Assists":        {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "Turnover":       {"base_low": -2, "base_high": -2, "threshold": 4},
    "Hockey Assist":  {"base_low": 2,  "base_high": 2,  "threshold": 4},
    "FT Made":        {"base_low": 2,  "base_high": 3,  "threshold": 4},
    "FT Missed":      {"base_low": -1, "base_high": -2, "threshold": 4},
}

DEFENSE_STATS: dict[str, dict] = {
    "Def. Reb":          {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "Steal":             {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "Deflection":        {"base_low": 3,  "base_high": 4,  "threshold": 4},
    "Def. Stop":         {"base_low": 3,  "base_high": 3,  "threshold": 4},
    "Charge":            {"base_low": 5,  "base_high": 7,  "threshold": 4},
    "Bluff":             {"base_low": 1,  "base_high": 1,  "threshold": 4},
    "Blocked Shot":      {"base_low": 2,  "base_high": 2,  "threshold": 4},
    "Jog Back":          {"base_low": -3, "base_high": -3, "threshold": 4},
    "No Ball Pressure":  {"base_low": -1, "base_high": -1, "threshold": 4},
    "Defensive Mistake": {"base_low": -1, "base_high": -1, "threshold": 4},
    "No Contest":        {"base_low": -1, "base_high": -1, "threshold": 4},
    "No Block Out":      {"base_low": -1, "base_high": -1, "threshold": 4},
    "Foul Against":      {"base_low": -1, "base_high": -1, "threshold": 4},
}

ALL_STAT_NAMES = list(OFFENSE_STATS.keys()) + list(DEFENSE_STATS.keys())


def stat_category(stat_name: str) -> str:
    """The only correct source of a stat's category.

    Every consumer — grades, quarter splits, the BIM composite, the prompt
    blocks — buckets rows by "offense"/"defense". A row stored under any other
    value is silently dropped from the off/def sums and raises a KeyError in the
    per-quarter split, so this must never be derived from the sign of the
    points: a missed shot is negative but still offense.
    """
    return "defense" if stat_name in DEFENSE_STATS else "offense"


def _quarter_multiplier(quarter: int) -> float:
    if quarter <= 2:
        return 1.0
    elif quarter == 3:
        return 1.25
    else:  # Q4 or OT
        return 1.5


# ── Post-game box-score import ─────────────────────────────────────────────────
# Mirror of the app's STAT_POINTS: stat_name -> (base_low, base_high, threshold).
_IMPORT_STAT_POINTS: dict[str, tuple[int, int, int]] = {
    "2 FG Made": (2, 3, 4), "3 FG Made": (3, 4, 4), "FT Made": (2, 3, 4),
    "Off. Reb": (3, 4, 4), "Def. Reb": (3, 4, 4), "Assists": (3, 4, 4),
    "Steal": (3, 4, 4), "Blocked Shot": (2, 2, 4), "Turnover": (-2, -2, 4),
    "Foul Against": (-1, -1, 4),
    # Misses. Absent from this table, an imported miss scored zero AND was
    # filtered out of the preview as an unknown stat — so a box score's
    # attempts had nowhere to land and no shooting percentage could be stated.
    # Same values live tracking uses, so an imported game and a tracked one
    # grade alike.
    "2 FG Missed": (-1, -2, 4), "3 FG Missed": (-1, -2, 4), "FT Missed": (-1, -2, 4),
}
# A box score covers the whole game, so it is stored under a neutral quarter
# with no clutch multiplier rather than being attributed to any one quarter.
IMPORT_QUARTER = 1
IMPORT_MULTIPLIER = 1.0


def _clear_prior_import(db: Session, game_id: int, is_opponent: bool | None = None) -> None:
    """Drop the previous imported box score for this game.

    Imports are whole-game totals, so importing twice would double every count.
    Only rows marked source="import" are removed — anything the coach tracked
    live during the game is left untouched.
    """
    q = db.query(models.GamePlayerStat).filter(
        models.GamePlayerStat.game_id == game_id,
        models.GamePlayerStat.source == "import",
    )
    if is_opponent is not None:
        q = q.filter(models.GamePlayerStat.is_opponent == is_opponent)
    q.delete(synchronize_session=False)


def _import_raw(stat: str, count: int) -> float:
    cfg = _IMPORT_STAT_POINTS.get(stat)
    if not cfg:
        return 0.0
    low, high, thr = cfg
    return float((high if count >= thr else low) * count)


BOX_SCORE_FIELDS = ("FGM", "FGA", "2PM", "2PA", "3PM", "3PA", "FTM", "FTA",
                    "OREB", "DREB", "REB", "AST", "STL", "BLK", "TO", "PF")

# Printed on every box score, and none of them a tally of events: minutes are a
# duration, a plus-minus is a difference and can be negative, efficiency is a
# formula. They are read off the sheet alongside the counted stats and stored
# on the player's line rather than as stat rows — see GameMinutesPlayed.
#
# Minutes are the one that changes a number the coach acts on: grading weights
# a player by time on the floor, and with none recorded everybody was treated
# as having played twenty.
LINE_FIELDS = ("MIN", "PM", "EFF")

_AI_BOX_SCORE_INSTRUCTION = (
    "This is a basketball box score. Return JSON only:\n"
    '{"teams": [{"team_name": "...", "players": [{"name": "...", '
    '"FGM": 0, "FGA": 0, "2PM": 0, "2PA": 0, "3PM": 0, "3PA": 0, "FTM": 0, "FTA": 0, '
    '"OREB": 0, "DREB": 0, "REB": 0, "AST": 0, "STL": 0, "BLK": 0, "TO": 0, "PF": 0, '
    '"MIN": "0:00", "PM": 0, "EFF": 0}]}]}\n'
    "Include every team and every player listed. Use the numbers exactly as printed — "
    "do NOT compute, estimate or fill in a stat that is not shown; omit the key instead. "
    "FGM/FGA are ALL field goals including threes. If the sheet shows a combined line "
    "like '9-18', FGM is 9 and FGA is 18. Skip team-total rows.\n"
    "MIN is minutes played exactly as printed — \"27:13\" stays \"27:13\", 27 stays 27. "
    "PM is the plus-minus column and CAN be negative; keep the sign. EFF is the "
    "efficiency column. Omit any of the three the sheet does not show rather than "
    "working it out."
)


def _canonical_rows_from_sheet(rows: list[list]) -> list[tuple[str, dict]]:
    """A spreadsheet's rows as (player, {FIELD: count}) — read exactly, no model."""
    import re as _re
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    name_idx = next((i for i, h in enumerate(header)
                     if h in ("name", "player", "player name", "athlete", "#")), 0)
    cols: dict[int, str] = {}
    for i, h in enumerate(header):
        for pat, field in _SHEET_COL_MAP:
            if _re.match(pat, h):
                cols[i] = field
                break
    out: list[tuple[str, dict]] = []
    for row in rows[1:]:
        if name_idx >= len(row) or not row[name_idx] or not str(row[name_idx]).strip():
            continue
        vals: dict[str, float] = {}
        for i, field in cols.items():
            if i >= len(row):
                continue
            if field == "MIN":
                mins = _minutes_value(row[i])
                if mins is not None:
                    vals[field] = mins
                continue
            if field in ("PM", "EFF"):
                # Kept even when negative: a minus-twelve is a result, and
                # dropping it left the column blank for the players it mattered
                # most for.
                signed = _signed_value(row[i])
                if signed is not None:
                    vals[field] = signed
                continue
            try:
                n = int(float(row[i]))
            except (TypeError, ValueError):
                continue
            if n >= 0:
                vals[field] = n
        out.append((str(row[name_idx]).strip(), vals))
    return out


# Header patterns → the canonical field names above. Kept in one table so the
# spreadsheet reader and the model's output end up in the same shape.
_SHEET_COL_MAP: list[tuple[str, str]] = [
    (r"^(3pm|3 ?fg ?made|3 ?pt ?made|3s ?made|threes ?made)", "3PM"),
    (r"^(3pa|3 ?fga|3 ?pt ?att|3s ?att|threes ?att)", "3PA"),
    (r"^(ftm|ft ?made|free ?throws? ?made)", "FTM"),
    (r"^(fta|ft ?att|free ?throws? ?att)", "FTA"),
    (r"^(2pm|2 ?fg ?made)", "2PM"),
    (r"^(2pa|2 ?fga)", "2PA"),
    (r"^(fgm|fg ?made)", "FGM"),
    (r"^(fga|fg ?att)", "FGA"),
    (r"^(oreb|o\.? ?reb|off\.? ?reb|offensive ?reb)", "OREB"),
    (r"^(dreb|d\.? ?reb|def\.? ?reb|defensive ?reb)", "DREB"),
    (r"^(reb|rebounds?|trb)", "REB"),
    (r"^(ast|assists?)", "AST"),
    (r"^(stl|steals?)", "STL"),
    (r"^(blk|blocks?)", "BLK"),
    (r"^(to|tov|turnovers?)", "TO"),
    (r"^(pf|fouls?)", "PF"),
    (r"^(min|mins|minutes|mp|time)", "MIN"),
    (r"^(\+/-|\+-|pm|plus ?minus|plusminus|\+/−)", "PM"),
    (r"^(eff|efficiency|effic|pir|index ?rating)", "EFF"),
]


def _minutes_value(raw) -> float | None:
    """Minutes, however the sheet writes them.

    "27:13" is twenty-seven minutes and thirteen seconds, not twenty-seven
    point thirteen — read as a decimal it would quietly under-count every
    player on the sheet.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text in ("-", "–", "—"):
        return None
    if ":" in text:
        parts = text.split(":")
        try:
            mins = int(parts[0] or 0)
            secs = int(parts[1] or 0) if len(parts) > 1 else 0
        except ValueError:
            return None
        return round(mins + secs / 60.0, 2)
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _signed_value(raw) -> float | None:
    """A number that is allowed to be negative, or None if the cell is blank."""
    if raw is None:
        return None
    text = str(raw).strip().replace("−", "-").replace("+", "")
    if not text or text in ("-", "–", "—"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _save_line_values(db: Session, game_id: int, is_opponent: bool,
                      player_name: str, vals: dict) -> None:
    """Minutes, plus-minus and efficiency for one player, from one sheet.

    Only what the sheet actually gave: a column it did not print stays None, so
    the app can say "not recorded" instead of showing a zero the coach would
    read as a real result.
    """
    mins = vals.get("MIN")
    pm = vals.get("PM")
    eff = vals.get("EFF")
    if mins is None and pm is None and eff is None:
        return
    row = (db.query(models.GameMinutesPlayed)
             .filter_by(game_id=game_id, player_name=player_name,
                        is_opponent=is_opponent).first())
    if row is None:
        row = models.GameMinutesPlayed(game_id=game_id, player_name=player_name,
                                       is_opponent=is_opponent, minutes_played=0.0)
        db.add(row)
    if mins is not None:
        row.minutes_played = float(mins)
    if pm is not None:
        row.plus_minus = float(pm)
    if eff is not None:
        row.efficiency = float(eff)


def _stats_from_canonical(v: dict) -> dict[str, int]:
    """Canonical box-score fields turned into BloomPrint's stat vocabulary.

    Two things this has to get right, both of which it used to get wrong:

    FGM is ALL field goals. Stored as "2 FG Made" it counted every three twice,
    so an imported 9-for-18 with three threes came out as 27 points, not 25.

    A miss is an attempt that was not a make. Attempts were never read at all,
    so the app knew the makes and could not state a single shooting percentage.
    Where a sheet gives no attempts the miss stays unrecorded rather than zero —
    zero would read as perfect shooting.
    """
    g = lambda k: max(int(v.get(k, 0) or 0), 0)
    threes, three_att = g("3PM"), g("3PA")
    twos = g("2PM") or max(g("FGM") - threes, 0)
    two_att = g("2PA") or max(g("FGA") - three_att, 0) if (g("2PA") or g("FGA")) else 0
    fts, ft_att = g("FTM"), g("FTA")

    out: dict[str, int] = {}
    def put(stat: str, n: int) -> None:
        if n > 0:
            out[stat] = out.get(stat, 0) + n

    put("2 FG Made", twos)
    put("3 FG Made", threes)
    put("FT Made", fts)
    if two_att:   put("2 FG Missed", max(two_att - twos, 0))
    if three_att: put("3 FG Missed", max(three_att - threes, 0))
    if ft_att:    put("FT Missed", max(ft_att - fts, 0))

    oreb, dreb, reb = g("OREB"), g("DREB"), g("REB")
    put("Off. Reb", oreb)
    # A sheet giving only a total: the offensive ones are already counted, so
    # the remainder is defensive rather than the total being double counted.
    put("Def. Reb", dreb or max(reb - oreb, 0))
    put("Assists", g("AST"))
    put("Steal", g("STL"))
    put("Blocked Shot", g("BLK"))
    put("Turnover", g("TO"))
    put("Foul Against", g("PF"))
    return out


# What a made shot is worth. Named once so the per-game path and the batched
# one cannot drift into scoring the same game differently.
_PER_POINT = {"2 FG Made": 2, "3 FG Made": 3, "FT Made": 1}


def derived_scores(game: models.GameSession) -> tuple[int | None, int | None]:
    """The final score worked out from the stats, when nobody typed one in.

    A game imported from a box score knows exactly what both teams scored — it
    is the sum of the made shots — but the score fields stay empty unless a
    coach fills them by hand. So the season record read 0W-0L with games in it,
    and the grade trend drew those games as losses, because "no score recorded"
    and "lost" look identical to code that only checks our_score > their_score.

    Both sides must have scoring stats. With numbers for one team only the
    honest answer is "unknown", not a shutout.
    """
    return scores_from_stats(game.player_stats)


def scores_from_stats(stats) -> tuple[int | None, int | None]:
    """The same sum, from stat rows already in hand.

    game.player_stats is a lazy relationship, so calling derived_scores() inside
    a loop over a season is one query per game to re-read rows the caller has
    usually just fetched.
    """
    made = {False: 0, True: 0}
    scored = {False: False, True: False}
    for st in stats:
        pts = _PER_POINT.get(st.stat_name)
        if pts is None:
            continue
        side = bool(st.is_opponent)
        made[side] += pts * (st.count or 0)
        scored[side] = True
    if not (scored[False] and scored[True]):
        return None, None
    return made[False], made[True]


def effective_scores(game: models.GameSession) -> tuple[int | None, int | None]:
    """What was typed in, or failing that what the stats say."""
    if game.our_score is not None and game.opponent_score is not None:
        return game.our_score, game.opponent_score
    return derived_scores(game)


def stats_need_reimport(db: Session, game_id: int) -> bool:
    """Was this game's box score imported before attempts were captured?

    An import that produced makes and no misses came from the old reader, which
    read no attempt columns at all. Two things follow and neither is visible:
    no shooting percentage is computable, and the points are inflated because
    FGM — all field goals — was stored as two-pointers, counting every three
    twice. Re-importing the same file corrects both.
    """
    rows = (db.query(models.GamePlayerStat)
              .filter(models.GamePlayerStat.game_id == game_id,
                      models.GamePlayerStat.source == "import").all())
    if not rows:
        return False
    return not any(r.stat_name.endswith("Missed") for r in rows)


def _side_for(team_name: str, ours: str, theirs: str) -> bool | None:
    """Which side a box score's team heading refers to — True for the opponent."""
    def norm(x: str) -> str:
        return "".join(ch for ch in (x or "").lower() if ch.isalnum())
    n, o, t = norm(team_name), norm(ours), norm(theirs)
    if not n:
        return None
    if o and (n in o or o in n):
        return False
    if t and (n in t or t in n):
        return True
    return None


@router.post("/sessions/{game_id}/import")
async def import_game_stats(
    game_id: int,
    files: list[UploadFile] = File(...),
    is_opponent: bool = False,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Import a post-game box score from any number of files, of any type.

    It used to take one .xlsx and reject everything else. A box score is far
    more often a PDF, a photo of the sheet taped to the scorer's table, or a
    screenshot — and a game's numbers are often spread over more than one of
    them. Anything readable goes now: spreadsheets are parsed exactly, and
    everything else is read by the same model that already reads roster photos.
    """
    import io as _io
    import openpyxl

    game = _get_game(db, game_id, coach.id)
    team_row = db.get(models.Team, game.team_id) if game.team_id else None
    our_name = (team_row.name if team_row else None) or coach.program_name or ""

    # (is_opponent, player, canonical stats) gathered across every file.
    collected: list[tuple[bool, str, dict]] = []
    read_errors: list[str] = []

    for f in files:
        content = await read_upload(f, what='document')
        name = (f.filename or "").lower()
        try:
            if name.endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
                rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
                if rows:
                    for player, vals in _canonical_rows_from_sheet(rows):
                        if vals:
                            collected.append((is_opponent, player, vals))
                    continue
            # Anything else — PDF, photo, screenshot, csv, Word — is read.
            data = ai_import.ai_extract_json(content, f.filename or "", f.content_type,
                                             _AI_BOX_SCORE_INSTRUCTION) or {}
            teams = data.get("teams") if isinstance(data, dict) else None
            for team in (teams or []):
                side = _side_for(str(team.get("team_name") or ""), our_name, game.opponent_name or "")
                for pl in (team.get("players") or []):
                    pname = str(pl.get("name") or "").strip()
                    if not pname:
                        continue
                    vals = {k: pl[k] for k in BOX_SCORE_FIELDS if isinstance(pl.get(k), (int, float))}
                    # MIN arrives as "27:13" as often as a number, and PM is
                    # signed — neither survives the numeric filter above.
                    mins = _minutes_value(pl.get("MIN"))
                    if mins is not None:
                        vals["MIN"] = mins
                    for key in ("PM", "EFF"):
                        signed = _signed_value(pl.get(key))
                        if signed is not None:
                            vals[key] = signed
                    if vals:
                        # A sheet naming neither team falls back to what the
                        # coach pressed, which is all the old importer ever had.
                        collected.append((is_opponent if side is None else side, pname, vals))
        except HTTPException:
            raise
        except Exception as exc:
            read_errors.append(f"{f.filename}: {exc}")

    if not collected:
        raise HTTPException(
            status_code=400,
            detail="No box score could be read from " +
                   ("those files." if len(files) > 1 else "that file.") +
                   (" " + "; ".join(read_errors) if read_errors else ""),
        )

    # A box score is whole-game totals, not a Q4 performance: recording it under
    # quarter 4 would apply the 1.5x clutch multiplier to the entire game and
    # inflate every imported game against live-tracked ones.
    q, mult = IMPORT_QUARTER, IMPORT_MULTIPLIER
    for side in {c[0] for c in collected}:
        _clear_prior_import(db, game.id, side)

    imported = 0
    for side, pname, vals in collected:
        _save_line_values(db, game.id, side, pname, vals)
        for stat, count in _stats_from_canonical(vals).items():
            raw = _import_raw(stat, count)
            db.add(models.GamePlayerStat(
                game_id=game.id, player_name=pname, is_opponent=side,
                quarter=q, stat_name=stat, stat_category=stat_category(stat),
                raw_points=raw, quarter_multiplier=mult, weighted_points=raw * mult, count=count,
                source="import",
            ))
            imported += 1
    db.commit()
    return {"imported": imported, "players": len(collected), "errors": read_errors}


def _compute_raw_points(stat_name: str, count: int) -> tuple[float, str]:
    """Returns (raw_points, category)."""
    if stat_name in OFFENSE_STATS:
        cfg = OFFENSE_STATS[stat_name]
        cat = "offense"
    elif stat_name in DEFENSE_STATS:
        cfg = DEFENSE_STATS[stat_name]
        cat = "defense"
    else:
        return (0.0, "offense")
    pv = cfg["base_high"] if count >= cfg["threshold"] else cfg["base_low"]
    return (float(pv * count), cat)


def player_tracked_stats_block(db, coach_id: int, player_name: str, game_ids) -> str:
    """Format a player's real tracked box-score stats for the given games into a
    prompt block. Shared by player summaries and new evals."""
    if not game_ids:
        return ""
    gstats = (
        db.query(models.GamePlayerStat)
        .join(models.GameSession, models.GameSession.id == models.GamePlayerStat.game_id)
        .filter(
            models.GameSession.coach_id == coach_id,
            models.GameSession.id.in_(list(game_ids)),
            models.GamePlayerStat.player_name == player_name,
            models.GamePlayerStat.is_opponent == False,
        )
        .all()
    )
    by_game: dict = {}
    for s in gstats:
        by_game.setdefault(s.game_id, []).append(s)
    lines = []
    for gid in game_ids:
        gs = by_game.get(gid)
        if not gs:
            continue
        game = db.get(models.GameSession, gid)
        counts: dict = {}
        for s in gs:
            counts[s.stat_name] = counts.get(s.stat_name, 0) + (s.count or 0)
        pts = counts.get("2 FG Made", 0) * 2 + counts.get("3 FG Made", 0) * 3 + counts.get("FT Made", 0)
        reb = counts.get("Off. Reb", 0) + counts.get("Def. Reb", 0)
        statline = ", ".join(f"{k}: {v}" for k, v in counts.items())
        when = game.date.strftime("%Y-%m-%d") if game and game.date else ""
        opp = game.opponent_name if game else "opponent"
        lines.append(
            f"vs {opp} ({when}) — PTS {pts}, REB {reb}, AST {counts.get('Assists', 0)}, "
            f"STL {counts.get('Steal', 0)}, BLK {counts.get('Blocked Shot', 0)}, TO {counts.get('Turnover', 0)}\n   {statline}"
        )
    if not lines:
        return ""
    return (
        "\n\nTRACKED GAME STATS (real box-score data from tracked games — build the box "
        "score from THESE numbers and reference them):\n" + "\n".join(lines)
    )


def _get_game(db: Session, game_id: int, coach_id: int) -> models.GameSession:
    game = db.get(models.GameSession, game_id)
    if not game or game.coach_id != coach_id:
        raise HTTPException(status_code=404, detail="Game session not found")
    return game


def _accessible_team_ids(db: Session, coach: models.Coach) -> set[int]:
    """Teams a coach can see games for: teams they own + teams they've joined
    as staff."""
    ids = {tm.id for tm in db.query(models.Team).filter_by(coach_id=coach.id).all()}
    ids |= {l.team_id for l in db.query(models.TeamStaff).filter_by(coach_id=coach.id).all()}
    return ids


def _coach_scouting(db: Session, coach: models.Coach, game: models.GameSession) -> str | None:
    """The coach's OWN scouting report for a game. Each coach who can access a
    game keeps their own private scouting; the owner's legacy report (stored on
    the game row before per-coach reports existed) is their own report."""
    row = (
        db.query(models.GameScoutingReport)
        .filter_by(game_id=game.id, coach_id=coach.id)
        .first()
    )
    if row and row.report_text:
        return row.report_text
    if game.coach_id == coach.id and game.ai_scouting_report:
        return game.ai_scouting_report
    return None


def _upsert_scouting(db: Session, coach: models.Coach, game: models.GameSession, text: str) -> None:
    row = (
        db.query(models.GameScoutingReport)
        .filter_by(game_id=game.id, coach_id=coach.id)
        .first()
    )
    if row:
        row.report_text = text
    else:
        db.add(models.GameScoutingReport(game_id=game.id, coach_id=coach.id, report_text=text))
    # Keep the legacy field in sync for the owner so older consumers still work.
    if game.coach_id == coach.id:
        game.ai_scouting_report = text


def _coach_game_report(db: Session, coach: models.Coach, game: models.GameSession) -> str | None:
    """The coach's OWN full game report for a game (per-coach, private)."""
    row = (
        db.query(models.GameFullReport)
        .filter_by(game_id=game.id, coach_id=coach.id)
        .first()
    )
    return row.report_text if row and row.report_text else None


def _upsert_game_report(db: Session, coach: models.Coach, game: models.GameSession, text: str) -> None:
    row = (
        db.query(models.GameFullReport)
        .filter_by(game_id=game.id, coach_id=coach.id)
        .first()
    )
    if row:
        row.report_text = text
    else:
        db.add(models.GameFullReport(game_id=game.id, coach_id=coach.id, report_text=text))


VERSIONS_KEPT = 20


# ── A recorded game, written out as a box score ────────────────────────────────
# The live tracker records events ("3 FG Made", "Def. Reb", "Charge"), not a box
# score. Those events ARE a box score once they are added up, which is what a
# coach means when they say "import the box score from that game" — they do not
# want the event log, they want PTS/FG/REB/AST the way it appears on a stat
# sheet. Anything BloomPrint tracks that a normal box score has no column for
# (deflections, defensive stops, charges, jog-backs) is kept underneath rather
# than dropped: it is often the most useful part, and it is why the game was
# tracked here rather than on paper.

def _pct(made: int, attempted: int) -> str:
    return f"{made}/{attempted}" + (f" ({round(100 * made / attempted)}%)" if attempted else "")


_BOX_SCORE_STATS = {
    "2 FG Made", "2 FG Missed", "3 FG Made", "3 FG Missed", "FT Made", "FT Missed",
    "Off. Reb", "Def. Reb", "Assists", "Steal", "Blocked Shot", "Turnover", "Foul Against",
}


def _side_box_score(stats: list, is_opp: bool) -> str:
    """One side of a game as a box-score table, plus whatever else was tracked."""
    counts: dict[str, dict[str, int]] = {}
    for s in stats:
        if bool(s.is_opponent) != is_opp:
            continue
        counts.setdefault(s.player_name, {})
        counts[s.player_name][s.stat_name] = (
            counts[s.player_name].get(s.stat_name, 0) + (s.count or 0))
    if not counts:
        return "(no tracked stats)\n"

    rows = []
    extras = []
    for name, c in sorted(counts.items()):
        g = c.get
        two_m, three_m, ft_m = g("2 FG Made", 0), g("3 FG Made", 0), g("FT Made", 0)
        fg_m = two_m + three_m
        fg_a = fg_m + g("2 FG Missed", 0) + g("3 FG Missed", 0)
        pts = two_m * 2 + three_m * 3 + ft_m
        rows.append([
            name, str(pts), _pct(fg_m, fg_a),
            _pct(three_m, three_m + g("3 FG Missed", 0)),
            _pct(ft_m, ft_m + g("FT Missed", 0)),
            str(g("Off. Reb", 0) + g("Def. Reb", 0)),
            str(g("Assists", 0)), str(g("Steal", 0)),
            str(g("Blocked Shot", 0)), str(g("Turnover", 0)),
            str(g("Foul Against", 0)),
        ])
        other = {k: v for k, v in c.items() if k not in _BOX_SCORE_STATS and v}
        if other:
            extras.append(f"- {name}: " + ", ".join(f"{k} {v}" for k, v in sorted(other.items())))

    head = ["Player", "PTS", "FG", "3PT", "FT", "REB", "AST", "STL", "BLK", "TO", "PF"]
    out = ("| " + " | ".join(head) + " |\n"
           + "| " + " | ".join("---" for _ in head) + " |\n"
           + "".join("| " + " | ".join(r) + " |\n" for r in rows))
    if extras:
        out += "\nAlso tracked:\n" + "\n".join(extras) + "\n"
    return out


def box_score_text(db: Session, game: models.GameSession) -> str:
    """A recorded game as text a coach can read and a report can be built from."""
    team = db.get(models.Team, game.team_id) if game.team_id else None
    coach = db.get(models.Coach, game.coach_id)
    our = (team.name if team else None) or (coach.program_name if coach else "My Team")
    when = game.date.strftime("%B %-d, %Y") if game.date else ""

    header = f"{our} vs {game.opponent_name}"
    if when:
        header += f" — {when}"
    if game.our_score is not None and game.opponent_score is not None:
        result = ("WIN" if game.our_score > game.opponent_score
                  else "LOSS" if game.our_score < game.opponent_score else "TIE")
        header += f"\nFinal: {our} {game.our_score}, {game.opponent_name} {game.opponent_score} ({result})"

    stats = list(game.player_stats)
    return (
        f"{header}\n\n"
        f"{our.upper()}:\n{_side_box_score(stats, False)}\n"
        f"{str(game.opponent_name).upper()}:\n{_side_box_score(stats, True)}"
    )


def _md_table(header: list[str], rows: list[list]) -> str:
    """A markdown pipe table — which the app and the PDF both draw as a grid."""
    if not rows:
        return ""
    line = lambda cells: "| " + " | ".join(str(c) for c in cells) + " |"
    return "\n".join([line(header), "|" + "|".join(["---"] * len(header)) + "|",
                       *[line(r) for r in rows]]) + "\n"


def _fmt_min(v) -> str:
    if v is None:
        return "-"
    mins = int(v)
    return f"{mins}:{int(round((v - mins) * 60)):02d}"


def _fmt_signed(v) -> str:
    if v is None:
        return "-"
    n = round(float(v), 1)
    n = int(n) if float(n).is_integer() else n
    return f"+{n}" if isinstance(n, (int, float)) and n > 0 else str(n)


def game_insights_text(db: Session, game: models.GameSession) -> str:
    """The Game Insights page, written out.

    Sharing a game used to send four lines — the opponent's name, the date, the
    score if somebody had typed one, and the scouting report if one existed. The
    numbers the game was actually about went nowhere, and the title named one
    team, so a staff member opened "vs Mali" with nothing in it.

    Everything here is already on the screen the sender was looking at. It is
    written as markdown tables because that is what the app and the exported PDF
    both draw as a grid.
    """
    team = db.get(models.Team, game.team_id) if game.team_id else None
    owner = db.get(models.Coach, game.coach_id)
    our = (team.name if team else None) or (owner.program_name if owner else "My Team")
    theirs = game.opponent_name or "Opponent"

    ours_pts, theirs_pts = effective_scores(game)
    out = [f"{our} vs {theirs}"]
    if game.date:
        out.append(game.date.strftime("%B %-d, %Y"))
    if ours_pts is not None and theirs_pts is not None:
        verdict = ("WIN" if ours_pts > theirs_pts
                   else "LOSS" if ours_pts < theirs_pts else "TIE")
        out.append(f"FINAL: {our} {ours_pts} — {theirs} {theirs_pts} ({verdict})")
    out.append("")

    stats = list(game.player_stats)
    extras = _line_extras(db, game.id)
    rosters = {False: _roster_jerseys(db, game.team_id),
               True: _roster_jerseys(db, _opponent_team_id(db, game))}
    sides = []
    for is_opp, name in ((False, our), (True, theirs)):
        rows, totals = _side_rows(stats, is_opp, rosters[is_opp], extras)
        sides.append((name, is_opp, rows, totals))

    # Team grades, the number at the top of the page.
    mp = db.query(models.GameMinutesPlayed).filter_by(game_id=game.id).all()
    mins = {False: {r.player_name: r.minutes_played for r in mp if not r.is_opponent},
            True: {r.player_name: r.minutes_played for r in mp if r.is_opponent}}
    grades = {}
    for _, is_opp, _, _ in sides:
        side_stats = [x for x in stats if bool(x.is_opponent) == is_opp]
        g = _compute_grades(side_stats, mins[is_opp])
        for_pts = ours_pts if not is_opp else theirs_pts
        against = theirs_pts if not is_opp else ours_pts
        grades[is_opp] = _team_grade(g, for_pts, against) if g else None
    if any(v is not None for v in grades.values()):
        out.append("TEAM GRADE")
        out.append(_md_table(["Team", "Grade"],
                             [[n, grades[o] if grades[o] is not None else "-"]
                              for n, o, _, _ in sides]))

    # Game leaders, over both teams, as the board shows them.
    everyone = [{**r, "team": n} for n, o, rows, _ in sides for r in rows]
    leader_rows = []
    for label, metric in (("Efficiency", _efficiency), ("Points", lambda r: r["PTS"]),
                          ("Rebounds", lambda r: r["REB"]), ("Assists", lambda r: r["AST"]),
                          ("Blocks", lambda r: r["BLK"]), ("Steals", lambda r: r["STL"])):
        best = max(everyone, key=metric, default=None)
        if best is not None and metric(best) > 0:
            leader_rows.append([label, best["player"], best["team"], metric(best)])
    if leader_rows:
        out.append("GAME LEADERS")
        out.append(_md_table(["", "Player", "Team", ""], leader_rows))

    # Shooting and the key stats, both teams side by side.
    def pct(m, a):
        return f"{round(100 * m / a, 1)}%" if a else "-"
    shoot = [[n, f"{t['FGM']}/{t['FGA']} {pct(t['FGM'], t['FGA'])}",
              f"{t['2PM']}/{t['2PA']} {pct(t['2PM'], t['2PA'])}",
              f"{t['3PM']}/{t['3PA']} {pct(t['3PM'], t['3PA'])}",
              f"{t['FTM']}/{t['FTA']} {pct(t['FTM'], t['FTA'])}"]
             for n, o, _, t in sides]
    out.append("SHOOTING")
    out.append(_md_table(["Team", "FG", "2PT", "3PT", "FT"], shoot))

    keys = ["PTS", "REB", "OREB", "DREB", "AST", "STL", "BLK", "TO", "PF"]
    out.append("KEY STATS")
    out.append(_md_table(["Team", *keys],
                         [[n, *[t.get(k, 0) for k in keys]] for n, o, _, t in sides]))

    # And both box scores in full, which is the part that was missing entirely.
    cols = ["MIN", "PTS", "FGM", "FGA", "3PM", "3PA", "FTM", "FTA",
            "OREB", "DREB", "REB", "AST", "STL", "BLK", "TO", "PF", "PM", "EFF"]
    labels = ["MIN", "PTS", "FGM", "FGA", "3PM", "3PA", "FTM", "FTA",
              "OREB", "DREB", "REB", "AST", "STL", "BLK", "TO", "PF", "+/-", "EFF"]
    for name, is_opp, rows, totals in sides:
        if not rows:
            continue
        body = []
        for r in rows:
            body.append([r["player"]] + [
                _fmt_min(r.get(c)) if c == "MIN"
                else _fmt_signed(r.get(c)) if c == "PM"
                else (r.get(c) if r.get(c) is not None else "-")
                for c in cols])
        body.append(["TOTAL"] + [
            _fmt_min(totals.get(c)) if c == "MIN"
            else _fmt_signed(totals.get(c)) if c == "PM"
            else (totals.get(c) if totals.get(c) is not None else "-")
            for c in cols])
        out.append(f"{name.upper()} — BOX SCORE")
        out.append(_md_table(["Player", *labels], body))

    return "\n".join(out).strip()


# ── The box score, and what can honestly be drawn from it ─────────────────────
#
# Everything here is derived from the stat events already recorded — tapped live
# or read out of an imported sheet. Nothing is estimated. Where the numbers to
# answer a question were never captured, the answer is absent rather than zero:
# a shooting percentage of 0% and "we never recorded the attempts" look nothing
# alike to a coach, and only one of them is true.

def _player_line(counts: dict[str, int]) -> dict:
    """One player's row: makes, attempts and the totals a box score prints."""
    g = lambda k: int(counts.get(k, 0) or 0)
    twos, threes, fts = g("2 FG Made"), g("3 FG Made"), g("FT Made")
    two_miss, three_miss, ft_miss = g("2 FG Missed"), g("3 FG Missed"), g("FT Missed")
    fgm, fga = twos + threes, twos + threes + two_miss + three_miss
    oreb, dreb = g("Off. Reb"), g("Def. Reb")
    return {
        "PTS": twos * 2 + threes * 3 + fts,
        "FGM": fgm, "FGA": fga,
        "2PM": twos, "2PA": twos + two_miss,
        "3PM": threes, "3PA": threes + three_miss,
        "FTM": fts, "FTA": fts + ft_miss,
        "OREB": oreb, "DREB": dreb, "REB": oreb + dreb,
        "AST": g("Assists"), "STL": g("Steal"), "BLK": g("Blocked Shot"),
        "TO": g("Turnover"), "PF": g("Foul Against"),
    }


_TOTALLED = ("PTS", "FGM", "FGA", "2PM", "2PA", "3PM", "3PA", "FTM", "FTA",
             "OREB", "DREB", "REB", "AST", "STL", "BLK", "TO", "PF")


def _line_extras(db: Session, game_id: int) -> dict[tuple[bool, str], dict]:
    """Minutes, plus-minus and efficiency per player, as the sheet printed them."""
    out: dict[tuple[bool, str], dict] = {}
    for r in db.query(models.GameMinutesPlayed).filter_by(game_id=game_id).all():
        out[(bool(r.is_opponent), r.player_name)] = {
            "MIN": r.minutes_played if r.minutes_played else None,
            "PM": r.plus_minus,
            "EFF": r.efficiency,
        }
    return out


def _side_rows(stats: list, is_opp: bool, roster: dict[str, str] | None = None,
               extras: dict[tuple[bool, str], dict] | None = None) -> tuple[list[dict], dict]:
    by_player: dict[str, dict[str, int]] = {}
    jerseys: dict[str, str] = {}
    for st in stats:
        if bool(st.is_opponent) != is_opp:
            continue
        d = by_player.setdefault(st.player_name, {})
        d[st.stat_name] = d.get(st.stat_name, 0) + (st.count or 0)
        # Any row that carries one will do; they all came off the same sheet.
        if getattr(st, "jersey_number", None) and st.player_name not in jerseys:
            jerseys[st.player_name] = str(st.jersey_number)
    rows = []
    for name, c in sorted(by_player.items()):
        row = {"player": name, "jersey": _jersey_for(name, jerseys, roster), **_player_line(c)}
        line = (extras or {}).get((is_opp, name), {})
        row["MIN"] = line.get("MIN")
        row["PM"] = line.get("PM")
        # The sheet's own efficiency where it printed one, and the standard
        # formula where it did not — the two agree, and a coach reading the
        # column should not have to know which they are looking at.
        row["EFF"] = line.get("EFF") if line.get("EFF") is not None else _efficiency(row)
        rows.append(row)
    rows.sort(key=lambda r: r["PTS"], reverse=True)
    totals = {k: sum(r[k] for r in rows) for k in _TOTALLED}
    # Minutes total to 200 on a five-a-side full game, which is how a coach
    # checks a sheet was read correctly.
    recorded_mins = [r["MIN"] for r in rows if r["MIN"] is not None]
    totals["MIN"] = round(sum(recorded_mins), 1) if recorded_mins else None
    pms = [r["PM"] for r in rows if r["PM"] is not None]
    totals["PM"] = round(sum(pms), 1) if pms else None
    totals["EFF"] = sum(r["EFF"] for r in rows if r["EFF"] is not None)
    return rows, totals


def _roster_jerseys(db: Session, team_id: int | None) -> dict[str, str]:
    """Squad numbers as the ROSTER has them, keyed by name.

    A number can arrive two ways — read off a stat sheet, or typed on the
    roster page — and only the first was ever displayed. A coach who filled one
    in by hand saw it on Roster and nowhere else in the app.
    """
    if not team_id:
        return {}
    return {_norm_team(p.name): str(p.jersey_number)
            for p in db.query(models.Player).filter_by(team_id=team_id).all()
            if p.jersey_number}


def _jersey_for(name: str, from_stats: dict[str, str], roster: dict[str, str] | None) -> str | None:
    """The roster first, then whatever the sheet said.

    The roster wins because it is the one a coach can edit: if they have typed a
    number against a name, that is the answer, and an import that disagrees has
    already written its own number there anyway.
    """
    if roster:
        at = roster.get(_norm_team(name))
        if at:
            return at
    return from_stats.get(name)


def _opponent_team_id(db: Session, game) -> int | None:
    """The team row for a game's opponent, when the coach keeps one."""
    if not game.opponent_name:
        return None
    row = next((tm for tm in db.query(models.Team).filter_by(coach_id=game.coach_id).all()
                if _norm_team(tm.name) == _norm_team(game.opponent_name)
                and tm.id != game.team_id), None)
    return row.id if row else None


def _efficiency(r: dict) -> int:
    """The standard efficiency line: what a player added minus what they cost."""
    return (r["PTS"] + r["REB"] + r["AST"] + r["STL"] + r["BLK"]
            - (r["FGA"] - r["FGM"]) - (r["FTA"] - r["FTM"]) - r["TO"])


def _lead_tracker(db: Session, game: models.GameSession) -> dict | None:
    """The game in order: who led, by how much, for how long.

    None of this is reachable from a box score. A total says the game finished
    80-71; only the sequence says one team led wire to wire, or that the lead
    changed nine times, or that an 11-point cushion evaporated in ninety
    seconds — which is the part a coach actually watches for.
    """
    # Period first: the sequence is only meaningful within the file it came
    # from, and halves can be imported on different days. An unlabelled event
    # has no place in the order, so it goes last rather than at the tip-off.
    events = (db.query(models.GamePlayEvent)
                .filter_by(game_id=game.id)
                .order_by(models.GamePlayEvent.period.is_(None),
                          models.GamePlayEvent.period,
                          models.GamePlayEvent.sequence).all())
    if not events:
        return None

    ours = theirs = 0
    points: list[dict] = []
    biggest_us = biggest_them = 0
    lead_changes = 0
    prev_leader = 0                 # -1 them, 0 tied, 1 us
    time_us = time_them = 0.0
    run_us = run_them = 0
    best_run_us = best_run_them = 0
    last_clock: float | None = None
    last_period: int | None = None

    for e in events:
        # A file that prints a running score is believed; otherwise it is added
        # up from the points on each event.
        if e.our_score is not None and e.opponent_score is not None:
            ours, theirs = e.our_score, e.opponent_score
        elif e.points:
            if e.is_opponent:
                theirs += e.points
            else:
                ours += e.points

        # Time spent in front, from the gap between consecutive clock readings.
        if e.clock_seconds is not None:
            if last_clock is not None and last_period == e.period:
                elapsed = max(last_clock - e.clock_seconds, 0.0)
                if prev_leader > 0:
                    time_us += elapsed
                elif prev_leader < 0:
                    time_them += elapsed
            last_clock, last_period = e.clock_seconds, e.period

        if e.points:
            if e.is_opponent:
                run_them += e.points
                run_us = 0
            else:
                run_us += e.points
                run_them = 0
            best_run_us = max(best_run_us, run_us)
            best_run_them = max(best_run_them, run_them)

        margin = ours - theirs
        biggest_us = max(biggest_us, margin)
        biggest_them = max(biggest_them, -margin)
        leader = 1 if margin > 0 else -1 if margin < 0 else 0
        # A tie is not a lead change; going from in front to behind is.
        if leader != 0 and prev_leader != 0 and leader != prev_leader:
            lead_changes += 1
        if leader != 0:
            prev_leader = leader

        points.append({"period": e.period, "clock": e.clock_seconds,
                       "our_score": ours, "opponent_score": theirs, "margin": margin})

    return {
        "points": points,
        "biggest_lead": {"us": biggest_us, "them": biggest_them},
        "biggest_run": {"us": best_run_us, "them": best_run_them},
        "time_leading": {"us": round(time_us), "them": round(time_them)},
        "lead_changes": lead_changes,
    }


def _advanced_stats(db: Session, game: models.GameSession) -> list[dict] | None:
    """The team-totals panel, as stated by whatever was imported.

    Never computed. Points off turnovers and the rest need possession context a
    box score does not carry, and a plausible guess at them is indistinguishable
    from the real number once it is drawn as a bar.
    """
    rows = db.query(models.GameTeamAdvanced).filter_by(game_id=game.id).all()
    fields = ("points_off_turnovers", "fast_break_points", "second_chance_points",
              "points_in_paint", "bench_points")
    # A totals panel that states PTS and REB but none of these is a Key Stats
    # panel, not this one. It used to count as "advanced stats exist", and since
    # every row draws nothing when neither side states it, the card came out as
    # two team names over empty space — which reads as a game where nobody did
    # any of it. Saying which file would fill it is the honest answer.
    if not any(getattr(r, f) is not None for r in rows for f in fields):
        return None
    return [{"is_opponent": bool(r.is_opponent),
             **{f: getattr(r, f) for f in fields}} for r in rows]


_PANEL_TO_BOX = {"PTS": "pts", "REB": "reb", "OREB": "oreb", "DREB": "dreb",
                 "AST": "ast", "STL": "stl", "BLK": "blk", "TO": "tov", "PF": "pf"}


def _official_totals(db: Session, game_id: int) -> dict[bool, dict]:
    """The team totals a sheet printed, by side.

    Preferred over the sum of the player rows wherever a sheet stated them: a
    box score credits team rebounds and team turnovers to the team, so its
    official REB and TO are legitimately higher than any set of named players
    adds up to. Summing the players left every team comparison a few short of
    the sheet the coach was holding.
    """
    out: dict[bool, dict] = {}
    for r in db.query(models.GameTeamAdvanced).filter_by(game_id=game_id).all():
        vals = {k: getattr(r, col) for k, col in _PANEL_TO_BOX.items() if getattr(r, col) is not None}
        if vals:
            out[bool(r.is_opponent)] = vals
    return out


def _shot_chart(db: Session, game: models.GameSession) -> list[dict] | None:
    shots = db.query(models.GameShot).filter_by(game_id=game.id).all()
    if not shots:
        return None
    return [{"is_opponent": bool(sh.is_opponent), "player": sh.player_name,
             "period": sh.period, "x": sh.x, "y": sh.y,
             "made": bool(sh.made), "points": sh.points} for sh in shots]


class PlayerLineIn(BaseModel):
    player_name: str
    is_opponent: bool = False
    # The same canonical fields the box score prints.
    line: dict


@router.put("/sessions/{game_id}/box-score/player")
def edit_player_line(
    game_id: int,
    body: PlayerLineIn,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Correct one player's line.

    A reader looking at a photograph of a stat sheet will occasionally put a 6
    where a 5 was, and there was nothing a coach could do about it but import
    the file again and hope. This replaces that player's imported rows with what
    the coach says they should be — only rows marked source="import", so
    anything tapped in live during the game is left alone.
    """
    game = _get_game(db, game_id, coach.id)
    name = (body.player_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Which player?")

    (db.query(models.GamePlayerStat)
       .filter(models.GamePlayerStat.game_id == game.id,
               models.GamePlayerStat.player_name == name,
               models.GamePlayerStat.is_opponent == bool(body.is_opponent),
               models.GamePlayerStat.source == "import")
       .delete(synchronize_session=False))

    line = {k: int(v) for k, v in (body.line or {}).items()
            if isinstance(v, (int, float)) and int(v) >= 0}
    for stat, count in _stats_from_canonical(line).items():
        raw = _import_raw(stat, count)
        db.add(models.GamePlayerStat(
            game_id=game.id, player_name=name, is_opponent=bool(body.is_opponent),
            quarter=IMPORT_QUARTER, stat_name=stat, stat_category=stat_category(stat),
            raw_points=raw, quarter_multiplier=IMPORT_MULTIPLIER,
            weighted_points=raw * IMPORT_MULTIPLIER, count=count, source="import",
        ))
    db.commit()
    return {"ok": True}


@router.get("/sessions/{game_id}/box-score")
def game_box_score(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """The game's numbers, and an honest statement of what is missing."""
    game = _get_game_readable(db, game_id, coach)
    team_row = db.get(models.Team, game.team_id) if game.team_id else None
    our_name = (team_row.name if team_row else None) or coach.program_name or "Us"
    stats = list(game.player_stats)

    official = _official_totals(db, game.id)
    # Both rosters, so a number typed on the Roster page shows up on the sheet.
    rosters = {False: _roster_jerseys(db, game.team_id),
               True: _roster_jerseys(db, _opponent_team_id(db, game))}
    extras = _line_extras(db, game.id)
    sides = []
    for is_opp, name in ((False, our_name), (True, game.opponent_name or "Opponent")):
        rows, totals = _side_rows(stats, is_opp, rosters[is_opp], extras)
        # The box score is the base, and the printed panel only fills in what the
        # player rows say nothing about.
        #
        # It used to be the other way round, to pick up team rebounds and team
        # turnovers, which belong to the team rather than to any player. But the
        # panel is a separate block on the sheet, often unlabelled, and when it
        # lands on the wrong side the coach is shown a Key Stats chart that flatly
        # contradicts the box score directly beneath it — 72 points above a column
        # of players adding to 83. A few team rebounds are not worth a total that
        # argues with the table it sits on. What the panel alone can say — points
        # off turnovers, fast break, second chance, paint, bench — is untouched.
        team_totals = dict(totals)
        for k, v in official.get(is_opp, {}).items():
            if not totals.get(k):
                team_totals[k] = v
        sides.append({"is_opponent": is_opp, "team_name": name,
                      "players": rows, "totals": team_totals,
                      "player_totals": totals,
                      "official_totals": bool(official.get(is_opp))})

    # Leaders, over both teams, so the board reads like a broadcast's.
    everyone = [{**r, "team_name": side["team_name"], "is_opponent": side["is_opponent"]}
                for side in sides for r in side["players"]]
    def top(metric) -> list[dict]:
        ranked = sorted(everyone, key=metric, reverse=True)[:5]
        return [{"player": r["player"], "team_name": r["team_name"], "value": metric(r)}
                for r in ranked if metric(r) > 0]
    leaders = {
        "efficiency": top(_efficiency),
        "points": top(lambda r: r["PTS"]),
        "rebounds": top(lambda r: r["REB"]),
        "assists": top(lambda r: r["AST"]),
        "blocks": top(lambda r: r["BLK"]),
        "steals": top(lambda r: r["STL"]),
    }

    # A percentage needs attempts. Where none were recorded there is no answer,
    # and saying 0% would be a lie the coach cannot see through.
    def pct(made: int, att: int):
        return round(100 * made / att, 1) if att > 0 else None
    shooting = [{
        "team_name": side["team_name"], "is_opponent": side["is_opponent"],
        "fg": pct(side["totals"]["FGM"], side["totals"]["FGA"]),
        "two": pct(side["totals"]["2PM"], side["totals"]["2PA"]),
        "three": pct(side["totals"]["3PM"], side["totals"]["3PA"]),
        "ft": pct(side["totals"]["FTM"], side["totals"]["FTA"]),
        "made": {k: side["totals"][k] for k in ("FGM", "FGA", "2PM", "2PA", "3PM", "3PA", "FTM", "FTA")},
    } for side in sides]

    has_attempts = any(s["totals"]["FGA"] or s["totals"]["FTA"] for s in sides)
    lead = _lead_tracker(db, game)
    advanced = _advanced_stats(db, game)
    shots = _shot_chart(db, game)
    return {
        "lead_tracker": lead,
        "advanced": advanced,
        "shot_chart": shots,
        "team_name": our_name,
        "opponent_name": game.opponent_name,
        "sides": sides,
        "leaders": leaders,
        "shooting": shooting,
        "key_stats": [k for k in ("PTS", "REB", "OREB", "DREB", "AST", "STL", "BLK", "TO", "PF")],
        # What can be drawn, and for what is missing, which file would supply it.
        "available": {
            "box_score": any(s["players"] for s in sides),
            "leaders": bool(everyone),
            "key_stats": bool(everyone),
            "shooting": has_attempts,
            "lead_tracker": lead is not None,
            "advanced": advanced is not None,
            "shot_chart": shots is not None,
        },
        "needs": {
            "shooting": None if has_attempts else "attempts",
            "lead_tracker": None if lead else "play_by_play",
            "advanced": None if advanced else "team_stats",
            "shot_chart": None if shots else "shot_coordinates",
        },
    }


@router.get("/sessions/{game_id}/box-score-text")
def get_box_score_text(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """The recorded game, written out as a box score."""
    game = _get_game_readable(db, game_id, coach)
    return {"text": box_score_text(db, game)}


def _game_team_ids(db: Session, coach, game) -> list[int | None]:
    """The teams a game is about: the one it was created under, and the
    opponent when they are a team on the books.

    Both, because a correction on a scouting report is usually about the
    opponent and a correction on a game report is usually about the coach's own
    side, and neither is worth guessing between.
    """
    ids: list[int | None] = [game.team_id]
    opp = _opponent_team_id_of(db, coach, game)
    if opp is not None and opp not in ids:
        ids.append(opp)
    # None is the program-wide scope, which applies to everything.
    ids.append(None)
    return ids


def _opponent_team_id_of(db: Session, coach, game) -> int | None:
    """The opponent's team row for this game, when they have one.

    A scouting report is about them, so a correction on it is about them too.
    None when the opponent is only ever a name on a fixture, which scopes the
    correction to the whole program instead — still true, just broader.
    """
    needle = (game.opponent_name or "").strip().lower()
    if not needle:
        return None
    for tm in db.query(models.Team).filter_by(coach_id=coach.id).all():
        if (tm.name or "").strip().lower() == needle:
            return tm.id
    return None


def team_written_material(db: Session, coach, team_name: str) -> list[tuple[str, str]]:
    """Everything WRITTEN about a team that Scout should be reading.

    The scouting page was built from games, box scores and per-game scouting
    reports, which means a coach could produce a full game packet about a team
    — with a film breakdown and a coaching report in it — and the page that
    exists to tell them about that team knew nothing about any of it.

    Returns [(what it is, the text)], newest first.
    """
    needle = "".join(ch for ch in (team_name or "").lower() if ch.isalnum())
    if not needle:
        return []

    def about(*names) -> bool:
        for n in names:
            flat = "".join(ch for ch in str(n or "").lower() if ch.isalnum())
            if flat and needle in flat:
                return True
        return False

    out: list[tuple[str, str]] = []
    packets = (db.query(models.GameReport)
                 .filter_by(coach_id=coach.id)
                 .order_by(models.GameReport.id.desc()).limit(40).all())
    for gr in packets:
        my_name = gr.my_team.name if gr.my_team else None
        opp_name = gr.opponent_team.name if gr.opponent_team else None
        if not about(my_name, opp_name, gr.opponent_name,
                     getattr(gr, "opponent_a_name", None), gr.title):
            continue
        if gr.report_text:
            out.append(("game packet report", gr.report_text))
        for clip in (gr.clips or []):
            # A clip names its own team; when it does not, the packet it sits
            # in is already known to be about this one.
            if clip.analysis_text and (not clip.team_name or about(clip.team_name)):
                out.append(("film breakdown", clip.analysis_text))
    return out


def team_material_block(db: Session, coach, team_name: str,
                        pieces: int = 3, chars: int = 3000) -> str:
    """The written material about a team, as prompt text.

    A scouting report was built from the numbers, the coach's notes, and film
    tied to that one game. The packet reports about the same team — the
    documents that say WHY those numbers happened — were not in it, so the app
    could hold a full coaching report on Angola and write a scouting report on
    Angola without reading a word of it.

    Capped rather than complete: a packet report is thousands of words and the
    newest ones carry the current picture.
    """
    written = team_written_material(db, coach, team_name)
    if not written:
        return ""
    parts = [f"\n\n[{kind}]\n{(text or '')[:chars]}" for kind, text in written[:pieces]]
    return ("\n\nREPORTS AND FILM ALREADY WRITTEN ABOUT THIS TEAM:"
            "\nUse what these already establish — schemes, personnel, tendencies — "
            "rather than working it out again from the numbers alone. Where they "
            "and the box score disagree, the box score is what happened."
            + "".join(parts))


def team_material_count(db: Session, coach, team_name: str) -> int:
    """How many written pieces exist about a team, for spotting staleness.

    A one-sentence insight used to be rewritten only when the GAME count
    changed, so a new report or packet about the team never touched it — the
    page kept showing a line written before the material it should have been
    written from.
    """
    return len(team_written_material(db, coach, team_name))


def learned_for_game(db: Session, coach, game) -> str:
    """What this coach's corrections have taught, for either team in this game.

    A correction is the coach saying what a report should be paying attention
    to, and they said it should hold for everything written about that team —
    which is this, not only the film breakdown it was made on.
    """
    from .preferences import for_prompt
    seen, out = set(), []
    for team_id in _game_team_ids(db, coach, game):
        block = for_prompt(db, coach.id, team_id)
        for line in block.splitlines():
            if line.strip().startswith("·") and line not in seen:
                seen.add(line)
                out.append(line)
    if not out:
        return ""
    return ("\n\nWHAT THIS COACH HAS ASKED FOR BEFORE:\n"
            "Corrections they have made to earlier reports. Treat them as standing "
            "instructions about what to look for and what to say — cover them where "
            "the evidence supports it, and do not claim them where it does not.\n"
            + "\n".join(out))


def film_notes_for_game(db: Session, coach, game) -> str:
    """Film breakdowns of THIS game, for a scouting or game report about it.

    A packet's film analysis and a tracked box score are two readings of the
    same night, and until they were linked a scouting report built from the
    numbers had no idea a film of that game had been watched at all — so it
    could describe what happened without ever saying what it looked like.

    Only films the coach has confirmed belong to this game. Nothing here is
    inferred from a name or a date.
    """
    clips = (db.query(models.GameReportClip)
               .filter(models.GameReportClip.game_id == game.id,
                       models.GameReportClip.analysis_text.isnot(None))
               .all())
    mine = []
    for clip in clips:
        packet = db.get(models.GameReport, clip.game_report_id)
        if packet and packet.coach_id == coach.id:
            mine.append(clip)
    if not mine:
        return ""
    out = ["\n\nFILM BREAKDOWNS OF THIS GAME:",
           "Analyses of film from this same game. Use what they saw — schemes, "
           "coverages, who did what — alongside the numbers below; they are "
           "describing the possessions the box score is counting."]
    for clip in mine:
        who = clip.team_name or ("our team" if clip.label == "my_team" else "the opponent")
        out.append(f"\n[Film — {who}]\n{(clip.analysis_text or '')[:6000]}")
    return "\n".join(out)


def _team_notes_text(db: Session, coach: models.Coach, game) -> str:
    """Everything the coach has asked to be remembered about EITHER team here.

    Notes are kept against a team's name, so a game between two named teams has
    two drawers to read. Reading only the opponent's meant a note a coach wrote
    about their own side went into the store and never came back out of it.
    """
    names: list[tuple[str, bool]] = []
    team_row = db.get(models.Team, game.team_id) if game.team_id else None
    ours = (team_row.name if team_row else None) or coach.program_name
    if ours:
        names.append((ours, False))
    if game.opponent_name and (not ours or ours.strip().lower() != game.opponent_name.strip().lower()):
        names.append((game.opponent_name, True))

    out = ""
    for name, is_opp in names:
        rows = (db.query(models.OpponentNote)
                  .filter_by(coach_id=coach.id, opponent_name=name)
                  .order_by(models.OpponentNote.created_at).all())
        if not rows:
            continue
        heading = f"COACH NOTES ON {name.upper()}" + (" (the opponent)" if is_opp else " (our team)")
        out += f"\n\n{heading}:\n" + "\n".join(f"- {n.note_text}" for n in rows)
    return out


def _file_edits(db: Session, game, coach: models.Coach, edits: list[str],
                remember: bool, correction_model, remember_team: str | None = None) -> None:
    """Keep what the coach typed, in whichever drawer they chose.

    The endpoint does this rather than the screen. When the screen owned it, a
    caller that did not happen to save the note first would have the text
    applied to the report and stored nowhere — it would shape this one
    regeneration and then be gone.

    A game has two teams and either can be the subject. `remember` files the
    note against the opponent; `remember_team` files it against a named team,
    which is how something worth keeping about the coach's own side is kept.
    Both may be on: notes are stored per team name, so the same sentence lands
    once for each.
    """
    for text in edits:
        kept = False
        if remember and game.opponent_name:
            db.add(models.OpponentNote(coach_id=coach.id,
                                       opponent_name=game.opponent_name, note_text=text))
            kept = True
        if remember_team:
            db.add(models.OpponentNote(coach_id=coach.id,
                                       opponent_name=remember_team, note_text=text))
            kept = True
        if not kept:
            db.add(correction_model(game_id=game.id, coach_id=coach.id, correction=text))
    if edits:
        db.commit()


def _edits_from_body(body: dict | None) -> list[str]:
    """The change the coach typed alongside the button they pressed."""
    if not body:
        return []
    text = (body.get("feedback") or body.get("text") or "").strip()
    return [text] if text else []


def _snapshot_report(db: Session, game_id: int, coach_id: int, kind: str, text: str) -> None:
    """Keep the wording about to be replaced.

    Only worth doing because regeneration now EDITS rather than rebuilds: while
    every regeneration was written fresh from the box score, a lost report could
    always be made again. An edited one cannot — the wording was the work.
    """
    if not (text or "").strip():
        return
    db.add(models.GameSessionReportVersion(
        game_id=game_id, coach_id=coach_id, kind=kind, report_text=text))
    db.flush()
    old = (db.query(models.GameSessionReportVersion)
             .filter_by(game_id=game_id, coach_id=coach_id, kind=kind)
             .order_by(models.GameSessionReportVersion.id.desc())
             .offset(VERSIONS_KEPT).all())
    for row in old:
        db.delete(row)


async def _apply_edits(prior_text: str, edits: list[str], coach: models.Coach,
                       on_words=None) -> str:
    """The report the coach already has, with the changes they asked for.

    The alternative — and what this replaced — was to write the report again
    from the box score and all accumulated context. That kept the report
    perfectly consistent with the numbers, and threw away everything about the
    existing one: a coach asking for one line to be corrected got a wholly
    different report back, with every observation reworded. Their context was
    never lost, but the report was.
    """
    from ..coach_context import language_directive

    prompt = (
        "Below is a basketball report, followed by changes the coach has asked for. "
        "Apply ALL of the changes.\n\n"
        "EDIT the report — do not rewrite it. Everything the changes do not touch must come "
        "back exactly as it is: same sections in the same order, same wording, same tables. "
        "Add what they ask to be added, remove what they ask to be removed, and correct what "
        "they say is wrong. Where a change contradicts the report, the coach saw the game and "
        "is right.\n\n"
        "Return ONLY the updated report, in the same format, with no preamble.\n\n"
        f"REPORT:\n{prior_text}\n\nCHANGES REQUESTED:\n"
        + "\n".join(f"- {e}" for e in edits)
        + f"{language_directive(coach)}"
    )
    return await long_text(prompt, max_tokens=12000, on_words=on_words)


def _scouting_text_for(db: Session, game_id: int, coach_id: int) -> str:
    """The scouting report this coach has on this game, for a finished job to
    hand back in the shape the screen already reads."""
    row = db.query(models.GameScoutingReport).filter_by(game_id=game_id, coach_id=coach_id).first()
    return row.report_text if row else ""


def _game_report_text_for(db: Session, game_id: int, coach_id: int) -> str:
    row = db.query(models.GameFullReport).filter_by(game_id=game_id, coach_id=coach_id).first()
    return row.report_text if row else ""


def _gate_context(db: Session, coach: models.Coach, games: list) -> dict:
    """Everything a list of games needs, in four queries instead of six per game.

    Serializing one game is cheap; serializing a season one game at a time is
    not. Each one asked for its own scouting row, its own full report, its own
    stat rows to work a score out of, and its own answer to "was this imported
    before attempts were captured" — which on a 24-game season is 76 round
    trips. Against a database on the other side of a network that is most of
    the wait before the Games page appears.
    """
    ids = [g.id for g in games]
    if not ids:
        return {"scouting": {}, "full": {}, "misses": set(), "imported": set(), "scores": {}}

    scouting = {r.game_id: r for r in db.query(models.GameScoutingReport)
                .filter(models.GameScoutingReport.game_id.in_(ids),
                        models.GameScoutingReport.coach_id == coach.id).all()}
    full = {r.game_id: r for r in db.query(models.GameFullReport)
            .filter(models.GameFullReport.game_id.in_(ids),
                    models.GameFullReport.coach_id == coach.id).all()}

    # Which games have imported rows at all, and which of those recorded a miss.
    # An import with makes and no misses came from the old reader — see
    # stats_need_reimport, which this replaces for the list case.
    imported, misses = set(), set()
    for gid, stat_name in (db.query(models.GamePlayerStat.game_id,
                                    models.GamePlayerStat.stat_name)
                             .filter(models.GamePlayerStat.game_id.in_(ids),
                                     models.GamePlayerStat.source == "import")
                             .distinct().all()):
        imported.add(gid)
        if str(stat_name).endswith("Missed"):
            misses.add(gid)

    # The score worked out from the made shots, for games nobody typed one on.
    made: dict[tuple[int, bool], int] = {}
    for gid, is_opp, stat_name, total in (
            db.query(models.GamePlayerStat.game_id, models.GamePlayerStat.is_opponent,
                     models.GamePlayerStat.stat_name, func.sum(models.GamePlayerStat.count))
              .filter(models.GamePlayerStat.game_id.in_(ids),
                      models.GamePlayerStat.stat_name.in_(_PER_POINT.keys()))
              .group_by(models.GamePlayerStat.game_id, models.GamePlayerStat.is_opponent,
                        models.GamePlayerStat.stat_name).all()):
        made[(gid, bool(is_opp))] = made.get((gid, bool(is_opp)), 0) \
            + _PER_POINT[stat_name] * int(total or 0)
    scores = {}
    for gid in ids:
        ours, theirs = made.get((gid, False)), made.get((gid, True))
        # Both sides or nothing: with numbers for one team only the honest
        # answer is "unknown", not a shutout. Same rule as derived_scores().
        scores[gid] = (ours, theirs) if ours is not None and theirs is not None else (None, None)
    return {"scouting": scouting, "full": full, "misses": misses,
            "imported": imported, "scores": scores}


def _gate_scouting(db: Session, coach: models.Coach, game: models.GameSession,
                   ctx: dict | None = None) -> schemas.GameSessionOut:
    """Serialize a game with the CURRENT coach's own scouting report. Staff see
    the game's stats/grades but only their own scouting write-up (empty until
    they generate one); scouting shared to them surfaces via Recent/Staff Hub.

    `ctx` is what _gate_context() prefetched for a whole list. Without it this
    answers every question itself, which is right for a single game.
    """
    out = schemas.GameSessionOut.model_validate(game)
    if ctx is not None:
        out.stats_need_reimport = game.id in ctx["imported"] and game.id not in ctx["misses"]
    else:
        out.stats_need_reimport = stats_need_reimport(db, game.id)
    if out.our_score is None or out.opponent_score is None:
        out.our_score, out.opponent_score = (
            ctx["scores"].get(game.id, (None, None)) if ctx is not None
            else effective_scores(game))

    row = ctx["scouting"].get(game.id) if ctx is not None else (
        db.query(models.GameScoutingReport)
          .filter_by(game_id=game.id, coach_id=coach.id).first())
    out.ai_scouting_report = (
        (row.report_text if row and row.report_text else None)
        or (game.ai_scouting_report if game.coach_id == coach.id else None))
    if out.ai_scouting_report:
        out.scouting_updated_at = (row.updated_at or row.created_at) if row else game.date

    grow = ctx["full"].get(game.id) if ctx is not None else (
        db.query(models.GameFullReport)
          .filter_by(game_id=game.id, coach_id=coach.id).first())
    # No legacy fallback here, unlike scouting: the full game report has always
    # been per-coach, so the row is the only place it lives.
    out.ai_game_report = grow.report_text if grow and grow.report_text else None
    if out.ai_game_report:
        out.game_report_updated_at = (grow.updated_at or grow.created_at) if grow else game.date
    return out


def _get_game_readable(db: Session, game_id: int, coach: models.Coach) -> models.GameSession:
    """Read access to a game: the owner, or any staff member linked to the
    game's team. Staff can view full game data (stats, grades, scouting) but
    editing stays owner-only via _get_game."""
    game = db.get(models.GameSession, game_id)
    if not game:
        raise HTTPException(status_code=404, detail="Game session not found")
    if game.coach_id == coach.id:
        return game
    if game.team_id is not None and game.team_id in _accessible_team_ids(db, coach):
        return game
    raise HTTPException(status_code=404, detail="Game session not found")


# ── Sessions ─────────────────────────────────────────────────────────────────

@router.post("/sessions", response_model=schemas.GameSessionOut)
def create_session(
    body: schemas.GameSessionCreate,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = models.GameSession(
        coach_id=coach.id,
        opponent_name=body.opponent_name,
        team_id=body.team_id,
        location=body.location,
        season_phase=body.season_phase,
        season_year=body.season_year,
        tracking_mode=body.tracking_mode,
        competition_level=body.competition_level,
        period_format=body.period_format or "quarters",
        num_periods=body.num_periods or 4,
        period_seconds=body.period_seconds or 480,
    )
    if body.date:
        try:
            game.date = datetime.fromisoformat(body.date)
        except ValueError:
            pass

    # Season-change nudge (activity-based) + activity stamp.
    from ..season import touch_and_maybe_remind
    game_year = (body.season_year or "").strip() or (str(game.date.year) if getattr(game, "date", None) else "")
    touch_and_maybe_remind(db, coach, game_year or None)

    db.add(game)
    db.commit()
    db.refresh(game)
    return game


@router.get("/sessions", response_model=list[schemas.GameSessionOut])
def list_sessions(
    season_phase: str | None = None,
    season_year: str | None = None,
    team_ids: str | None = None,  # comma-separated; omitted means every team
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    # A coach sees their own games plus games on any team they're linked to as
    # staff, so a team's schedule shows up in the Team Grade tab for all staff.
    from sqlalchemy import or_
    # Named apart from the team_ids QUERY PARAMETER above. When both were
    # called team_ids the local won, the filter was handed a list of ints, and
    # it silently parsed to "no selection" — the picker appeared to work and
    # changed nothing.
    accessible_ids = _accessible_team_ids(db, coach)
    conds = [models.GameSession.coach_id == coach.id]
    if accessible_ids:
        conds.append(models.GameSession.team_id.in_(accessible_ids))
    q = db.query(models.GameSession).filter(or_(*conds))
    picked = _team_filter(db, team_ids)
    if picked is not None:
        q = q.filter(picked)
    if season_phase:
        q = q.filter(models.GameSession.season_phase == season_phase)
    if season_year:
        q = q.filter(models.GameSession.season_year == season_year)
    games = q.order_by(models.GameSession.date.desc()).all()
    # Gate each game's written scouting report: staff see the game's stats but
    # only see the AI scouting write-up if it was shared to them.
    ctx = _gate_context(db, coach, games)
    return [_gate_scouting(db, coach, g, ctx) for g in games]


@router.get("/sessions/{game_id}", response_model=schemas.GameSessionOut)
def get_session(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    return _gate_scouting(db, coach, _get_game_readable(db, game_id, coach))


@router.patch("/sessions/{game_id}", response_model=schemas.GameSessionOut)
def update_session(
    game_id: int,
    body: schemas.GameSessionUpdate,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(game, field, value)
    db.commit()
    db.refresh(game)
    return game


@router.delete("/sessions/{game_id}")
def delete_session(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    # Hidden, not destroyed — a tracked game carries a whole night's stats.
    soft_delete(db, game)
    db.commit()
    return {"ok": True}


# ── Stats ─────────────────────────────────────────────────────────────────────

@router.get("/sessions/{game_id}/stats")
def list_stats(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game_readable(db, game_id, coach)
    stats = (
        db.query(models.GamePlayerStat)
        .filter_by(game_id=game.id)
        .order_by(models.GamePlayerStat.id)
        .all()
    )
    return [
        {
            "id": s.id,
            "player_name": s.player_name,
            "is_opponent": s.is_opponent,
            "quarter": s.quarter,
            "stat_name": s.stat_name,
            "stat_category": s.stat_category,
            "raw_points": s.raw_points,
            "weighted_points": s.weighted_points,
            "count": s.count,
        }
        for s in stats
    ]


@router.post("/sessions/{game_id}/stats")
def log_stat(
    game_id: int,
    body: schemas.GameStatEntry,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    multiplier = _quarter_multiplier(body.quarter)
    weighted = body.raw_points * multiplier
    stat = models.GamePlayerStat(
        game_id=game.id,
        player_id=body.player_id,
        player_name=body.player_name,
        is_opponent=body.is_opponent,
        quarter=body.quarter,
        stat_name=body.stat_name,
        stat_category=body.stat_category,
        raw_points=body.raw_points,
        quarter_multiplier=multiplier,
        weighted_points=weighted,
        count=body.count,
    )
    db.add(stat)
    db.commit()
    db.refresh(stat)
    return {"id": stat.id, "weighted_points": weighted}


@router.delete("/stats/{stat_id}")
def delete_stat(
    stat_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    stat = db.get(models.GamePlayerStat, stat_id)
    if not stat:
        raise HTTPException(status_code=404, detail="Stat not found")
    # Verify ownership via game
    game = db.get(models.GameSession, stat.game_id)
    if not game or game.coach_id != coach.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    db.delete(stat)
    db.commit()
    return {"ok": True}


# ── Lineup ────────────────────────────────────────────────────────────────────

@router.post("/sessions/{game_id}/lineup")
def log_lineup(
    game_id: int,
    body: schemas.LineupEventCreate,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    event = models.LineupEvent(
        game_id=game.id,
        player_id=body.player_id,
        player_name=body.player_name,
        is_opponent=body.is_opponent,
        event_type=body.event_type,
        quarter=body.quarter,
        timestamp_seconds=body.timestamp_seconds,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    return {"id": event.id}


@router.get("/sessions/{game_id}/lineup")
def get_lineup(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game_readable(db, game_id, coach)
    events = (
        db.query(models.LineupEvent)
        .filter_by(game_id=game.id)
        .order_by(models.LineupEvent.created_at)
        .all()
    )
    return [
        {
            "id": e.id,
            "player_name": e.player_name,
            "is_opponent": e.is_opponent,
            "event_type": e.event_type,
            "quarter": e.quarter,
            "timestamp_seconds": e.timestamp_seconds,
        }
        for e in events
    ]


# ── Minutes ───────────────────────────────────────────────────────────────────

@router.post("/sessions/{game_id}/minutes")
def log_minutes(
    game_id: int,
    body: schemas.GameMinutesEntry,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    existing = (
        db.query(models.GameMinutesPlayed)
        .filter_by(game_id=game.id, player_name=body.player_name, is_opponent=body.is_opponent)
        .first()
    )
    if existing:
        existing.minutes_played = body.minutes_played
    else:
        mp = models.GameMinutesPlayed(
            game_id=game.id,
            player_id=body.player_id,
            player_name=body.player_name,
            is_opponent=body.is_opponent,
            minutes_played=body.minutes_played,
        )
        db.add(mp)
    db.commit()
    return {"ok": True}


# ── Summary ───────────────────────────────────────────────────────────────────

def _compute_grades(stats: list[models.GamePlayerStat], minutes_map: dict[str, float],
                    line_extras: dict[str, dict] | None = None) -> list[dict]:
    player_data: dict[str, dict] = {}
    for s in stats:
        if s.player_name not in player_data:
            player_data[s.player_name] = {
                "offensive_weighted": 0.0,
                "defensive_weighted": 0.0,
                "jersey_number": None,
                "quarters": defaultdict(lambda: {"offense": 0.0, "defense": 0.0}),
            }
        pd = player_data[s.player_name]
        pd["jersey_number"] = pd["jersey_number"] or getattr(s, "jersey_number", None)
        if s.stat_category == "offense":
            pd["offensive_weighted"] += s.weighted_points
        else:
            pd["defensive_weighted"] += s.weighted_points
        pd["quarters"][s.quarter][s.stat_category] += s.weighted_points

    grades = []
    for name, data in player_data.items():
        # Real minutes when the sheet gave them. Where it did not, the grade
        # is still worked out the way it always was — blanking every historical
        # game would be worse than the fault being fixed — but the minutes
        # themselves are reported as unknown rather than as twenty, so nothing
        # downstream prints a number nobody recorded.
        extras = (line_extras or {}).get(name, {})
        recorded = minutes_map.get(name)
        mins = recorded if recorded and recorded > 0 else 20.0
        total = data["offensive_weighted"] + data["defensive_weighted"]
        game_grade = round(total / max(mins, 1.0), 2)
        grades.append({
            "player_name": name,
            "jersey_number": data.get("jersey_number"),
            "offensive_grade": round(data["offensive_weighted"], 2),
            "defensive_grade": round(data["defensive_weighted"], 2),
            "total_grade": round(total, 2),
            "minutes_played": round(recorded, 1) if recorded else None,
            "minutes_recorded": recorded is not None,
            "game_grade": game_grade,
            "plus_minus": extras.get("plus_minus"),
            "efficiency": extras.get("efficiency"),
            "per_quarter": {str(q): dict(v) for q, v in data["quarters"].items()},
        })
    grades.sort(key=lambda x: x["game_grade"], reverse=True)
    return grades


def _win_loss_factor(for_score: int | None, against_score: int | None) -> float:
    """The 40% of a team grade that is about the result, from one side's view.

    Written from a point of view rather than hard-coded to "us" so the same
    scale can be applied to the opponent: a team that won by twenty and a team
    that lost by twenty must not both come out at 5.0.
    """
    if for_score is None or against_score is None:
        return 5.0
    if for_score > against_score:
        return 10.0
    return 7.0 if abs(for_score - against_score) <= 5 else 5.0


def _team_grade(grades: list[dict], for_score, against_score) -> float:
    avg = sum(g["game_grade"] for g in grades) / len(grades) if grades else 0.0
    return round((avg * 0.6) + (_win_loss_factor(for_score, against_score) * 0.4), 2)


@router.get("/sessions/{game_id}/summary")
def get_summary(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game_readable(db, game_id, coach)

    # Minutes map
    mp_records = db.query(models.GameMinutesPlayed).filter_by(game_id=game.id).all()
    our_minutes = {r.player_name: r.minutes_played for r in mp_records if not r.is_opponent}
    opp_minutes = {r.player_name: r.minutes_played for r in mp_records if r.is_opponent}
    # The plus-minus and the efficiency off the same sheet, so a grade row can
    # show them beside the grade instead of a hardcoded zero.
    our_extras = {r.player_name: {"plus_minus": r.plus_minus, "efficiency": r.efficiency}
                  for r in mp_records if not r.is_opponent}
    opp_extras = {r.player_name: {"plus_minus": r.plus_minus, "efficiency": r.efficiency}
                  for r in mp_records if r.is_opponent}

    our_stats = [s for s in game.player_stats if not s.is_opponent]
    opp_stats = [s for s in game.player_stats if s.is_opponent]

    player_grades = _compute_grades(our_stats, our_minutes, our_extras)
    opponent_grades = _compute_grades(opp_stats, opp_minutes, opp_extras)
    # A squad number belongs to the player, not to the sheet they were read
    # from — so the roster answers first wherever a grade is shown.
    for grades, team_id in ((player_grades, game.team_id),
                            (opponent_grades, _opponent_team_id(db, game))):
        roster = _roster_jerseys(db, team_id)
        for g in grades:
            g["jersey_number"] = _jersey_for(
                g["player_name"],
                {g["player_name"]: g["jersey_number"]} if g.get("jersey_number") else {},
                roster)

    ours, theirs = effective_scores(game)
    team_grade = _team_grade(player_grades, ours, theirs)
    # The same grade from the other bench. With both teams' stats in there is
    # no reason to grade one of them: the opponent's number is what a scouting
    # report is arguing with, and one grade on a page about two teams reads as
    # if the other side did not play.
    opponent_grade = _team_grade(opponent_grades, theirs, ours) if opponent_grades else None

    game_out = _gate_scouting(db, coach, game)
    return {
        "game": game_out.model_dump(),
        "player_grades": player_grades,
        "team_grade": team_grade,
        "opponent_grades": opponent_grades,
        "opponent_team_grade": opponent_grade,
    }


@router.get("/player-game-history")
def player_game_history(
    player_name: str,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """All games this player appeared in, with per-game stats and grade.
    Covers the coach's own games plus games on teams they're staff on, so the
    Team Grade leaderboard detail works for shared team schedules."""
    from sqlalchemy import or_
    team_ids = _accessible_team_ids(db, coach)
    game_conds = [models.GameSession.coach_id == coach.id]
    if team_ids:
        game_conds.append(models.GameSession.team_id.in_(team_ids))
    stats = (
        db.query(models.GamePlayerStat)
        .join(models.GameSession, models.GameSession.id == models.GamePlayerStat.game_id)
        .filter(
            or_(*game_conds),
            models.GamePlayerStat.player_name == player_name,
            models.GamePlayerStat.is_opponent == False,
        )
        .all()
    )

    game_ids = sorted({s.game_id for s in stats}, reverse=True)
    result = []
    for game_id in game_ids:
        game = db.get(models.GameSession, game_id)
        if not game:
            continue
        gs = [s for s in stats if s.game_id == game_id]

        off_w = sum(s.weighted_points for s in gs if s.stat_category == "offense")
        def_w = sum(s.weighted_points for s in gs if s.stat_category == "defense")
        total_w = off_w + def_w

        mp_rec = (
            db.query(models.GameMinutesPlayed)
            .filter_by(game_id=game_id, player_name=player_name, is_opponent=False)
            .first()
        )
        recorded_minutes = mp_rec.minutes_played if mp_rec and mp_rec.minutes_played else None
        minutes = recorded_minutes or 20.0
        game_grade = round(total_w / max(minutes, 1.0), 2)

        stat_breakdown: dict = {}
        for s in gs:
            if s.stat_name not in stat_breakdown:
                stat_breakdown[s.stat_name] = {"count": 0, "weighted_points": 0.0, "category": s.stat_category}
            stat_breakdown[s.stat_name]["count"] += s.count
            stat_breakdown[s.stat_name]["weighted_points"] += round(s.weighted_points, 2)

        quarters: dict = {}
        for s in gs:
            q = str(s.quarter)
            if q not in quarters:
                quarters[q] = {"offense": 0.0, "defense": 0.0, "stats": []}
            quarters[q][s.stat_category] = round(quarters[q][s.stat_category] + s.weighted_points, 2)
            quarters[q]["stats"].append({
                "name": s.stat_name,
                "count": s.count,
                "weighted_points": round(s.weighted_points, 2),
                "category": s.stat_category,
            })

        result.append({
            "game_id": game.id,
            "opponent_name": game.opponent_name,
            "date": game.date.strftime("%B %d, %Y") if game.date else None,
            "year": game.date.year if game.date else None,
            "season_year": game.season_year,
            "our_score": game.our_score,
            "opponent_score": game.opponent_score,
            "season_phase": game.season_phase,
            "game_grade": game_grade,
            "offensive_weighted": round(off_w, 2),
            "defensive_weighted": round(def_w, 2),
            "minutes": round(minutes, 1),
            "stat_breakdown": stat_breakdown,
            "per_quarter": quarters,
        })

    return result


# ── Upload Excel ──────────────────────────────────────────────────────────────

@router.post("/sessions/{game_id}/upload")
async def upload_excel(
    game_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game = _get_game(db, game_id, coach.id)
    content = await read_upload(file, what='spreadsheet')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    _clear_prior_import(db, game.id)
    imported = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row with player names
        header_row_idx = None
        player_cols: dict[int, str] = {}  # col_idx -> player_name
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if cell and "Game Grade" in str(cell):
                    header_row_idx = i
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            # Try first row as header
            header_row_idx = 0

        header_row = rows[header_row_idx]
        for j, cell in enumerate(header_row):
            val = str(cell).strip() if cell else ""
            if val and val not in ("", "None", "Stat", "Game Grade Sheet"):
                player_cols[j] = val

        # Parse stat rows
        for row in rows[header_row_idx + 1:]:
            if not row or not row[0]:
                continue
            stat_raw = str(row[0]).strip()
            if not stat_raw:
                continue

            # Fuzzy match stat name
            matches = difflib.get_close_matches(stat_raw, ALL_STAT_NAMES, n=1, cutoff=0.6)
            if not matches:
                continue
            stat_name = matches[0]
            cat = stat_category(stat_name)

            for col_idx, player_name in player_cols.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                try:
                    count = int(float(str(val)))
                except (ValueError, TypeError):
                    continue
                if count == 0:
                    continue

                raw_points, _ = _compute_raw_points(stat_name, count)
                weighted = raw_points * IMPORT_MULTIPLIER

                stat = models.GamePlayerStat(
                    game_id=game.id,
                    player_name=player_name,
                    is_opponent="opp" in sheet_name.lower() or "opponent" in sheet_name.lower(),
                    quarter=IMPORT_QUARTER,
                    stat_name=stat_name,
                    stat_category=cat,
                    raw_points=raw_points,
                    quarter_multiplier=IMPORT_MULTIPLIER,
                    weighted_points=weighted,
                    count=count,
                    source="import",
                )
                db.add(stat)
                imported += 1

    db.commit()
    return {"imported": imported}


# ── AI Scouting Report ────────────────────────────────────────────────────────

async def _run_scouting(db: Session, coach: models.Coach, game: models.GameSession,
                        corrections: list[str], on_words=None) -> str:
    """Build + run the scouting report from the box score plus any coach-added
    context ('corrections'), upsert it for this coach, and return the text."""
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not configured.")

    opp_stats = [s for s in game.player_stats if s.is_opponent]
    opp_summary: dict[str, dict] = {}
    for s in opp_stats:
        if s.player_name not in opp_summary:
            opp_summary[s.player_name] = {"offense": 0.0, "defense": 0.0, "stats": defaultdict(int)}
        opp_summary[s.player_name][s.stat_category] += s.weighted_points
        opp_summary[s.player_name]["stats"][s.stat_name] += s.count

    opp_context = ""
    for pname, data in opp_summary.items():
        top_stats = sorted(data["stats"].items(), key=lambda x: x[1], reverse=True)[:5]
        opp_context += f"\n{pname}: OFF={data['offense']:.1f} DEF={data['defense']:.1f}  Top stats: {', '.join(f'{s}={c}' for s,c in top_stats)}"

    score_info = ""
    if game.our_score is not None and game.opponent_score is not None:
        result = "WIN" if game.our_score > game.opponent_score else "LOSS"
        score_info = f"Final score: {game.our_score}-{game.opponent_score} ({result})"

    notes_text = (_team_notes_text(db, coach, game) + film_notes_for_game(db, coach, game)
                  + team_material_block(db, coach, game.opponent_name or "")
                  + learned_for_game(db, coach, game))

    corr_text = ""
    if corrections:
        corr_text = (
            "\n\nCOACH CONTEXT & ADJUSTMENTS (qualitative detail the box score can't capture — "
            "you MUST weave ALL of these into the analysis and recommendations):\n"
            + "\n".join(f"- {c}" for c in corrections if c and c.strip())
        )

    from ..coach_context import resolve_level, language_directive
    _lvl = resolve_level(coach, team=(db.get(models.Team, game.team_id) if game.team_id else None))
    prompt = (
        f"You are the BloomPrint Basketball Intelligence Model. "
        f"Generate a pre-game scouting report for the opponent: {game.opponent_name}\n\n"
        f"Game date: {game.date}\n"
        f"COMPETITION LEVEL: {_lvl} — calibrate every read, tendency, and recommendation to this level.\n"
        f"{score_info}\n\n"
        f"OPPONENT PLAYER GRADES:\n{opp_context}"
        f"{notes_text}"
        f"{corr_text}\n\n"
        f"Analyze the opponent's strengths, weaknesses, top players to watch, offensive tendencies, "
        f"defensive tendencies, and strategic recommendations for the next game against them."
        f"{REPORT_FORMAT}"
        f"{language_directive(coach)}"
    )

    try:
        # Streamed. A non-streaming request carries a ten-minute SDK ceiling and
        # two silent retries behind it, so a report that ran long was abandoned
        # and paid for three times — see ai_models.long_text.
        report_text = await long_text(prompt, max_tokens=8000, on_words=on_words)
        if not report_text.strip():
            raise HTTPException(status_code=500, detail="AI returned no content")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {exc}")

    import re as _re
    report_text = _re.sub(r"\s*END OF REPORT\.?\s*$", "", report_text, flags=_re.IGNORECASE).rstrip()
    _upsert_scouting(db, coach, game, report_text)
    db.commit()
    return report_text


def _run_scouting_job(game_id: int, coach_id: int, job_id: int,
                      extra_edits: list[str] | None = None) -> None:
    """The scouting report, off the request that asked for it. See api/genjob.py.

    Always written from ALL of the coach's added context, and marks whatever was
    still pending as applied — which is what the "Apply & Regenerate" button did
    and what the first generation does with an empty list.
    """
    import asyncio

    def work():
        db = SessionLocal()
        try:
            game = db.get(models.GameSession, game_id)
            coach = db.get(models.Coach, coach_id)
            if not game or not coach:
                raise RuntimeError("Game not found")
            rows = (db.query(models.GameScoutingCorrection)
                      .filter_by(game_id=game_id, coach_id=coach_id)
                      .order_by(models.GameScoutingCorrection.id).all())
            prior = _scouting_text_for(db, game_id, coach_id)
            _snapshot_report(db, game_id, coach_id, "scouting", prior)
            edits = list(extra_edits or []) + [c.correction for c in rows if not c.applied]
            if prior and edits:
                text = asyncio.run(_apply_edits(prior, edits, coach,
                                                on_words=genjob.words_reporter(job_id)))
                _upsert_scouting(db, coach, game, text.strip())
            else:
                # No report to edit, or nothing new to apply — write it from the
                # box score and every piece of context on file.
                asyncio.run(_run_scouting(db, coach, game, [c.correction for c in rows],
                                          on_words=genjob.words_reporter(job_id)))
            for r in rows:
                r.applied = True
            db.commit()
        finally:
            db.close()
        return game_id

    genjob.run(job_id, work)


@router.get("/sessions/{game_id}/report-versions")
def list_report_versions(
    game_id: int,
    kind: str = "scouting",
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Previous wordings of this coach's report for this game, newest first."""
    _get_game_readable(db, game_id, coach)
    rows = (db.query(models.GameSessionReportVersion)
              .filter_by(game_id=game_id, coach_id=coach.id, kind=kind)
              .order_by(models.GameSessionReportVersion.id.desc()).all())
    return [{"id": r.id, "created_at": r.created_at, "report_text": r.report_text}
            for r in rows]


@router.post("/sessions/{game_id}/report-versions/{version_id}/restore")
def restore_report_version(
    game_id: int,
    version_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Put a previous wording back, keeping the current one as a version so
    restoring is itself reversible."""
    game = _get_game_readable(db, game_id, coach)
    row = db.get(models.GameSessionReportVersion, version_id)
    if not row or row.game_id != game_id or row.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="That version was not found.")
    current = (_scouting_text_for if row.kind == "scouting" else _game_report_text_for)(
        db, game_id, coach.id)
    _snapshot_report(db, game_id, coach.id, row.kind, current)
    if row.kind == "scouting":
        _upsert_scouting(db, coach, game, row.report_text or "")
        db.commit()
        return {"ai_scouting_report": row.report_text}
    _upsert_game_report(db, coach, game, row.report_text or "")
    db.commit()
    return {"report_text": row.report_text, "ai_game_report": row.report_text}


@router.post("/sessions/{game_id}/ai-scouting-job")
def start_ai_scouting(
    game_id: int,
    background_tasks: BackgroundTasks,
    body: dict | None = None,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Start the scouting report and hand back a job to follow.

    Serves both the first generation and "Apply & Regenerate": an optional
    feedback body is recorded as context before the report is written.
    """
    game = _get_game_readable(db, game_id, coach)
    # What the coach just typed is applied to the report whichever drawer it is
    # filed in. "Remember for this opponent" decides where the text is KEPT —
    # durable opponent note or one-off correction for this report — and used to
    # decide, as a side effect, whether it reached the report at all: a
    # remembered note was not a correction, so nothing was pending, so the
    # report was rebuilt and the note only landed because a rebuild reads notes
    # too. Once regeneration edits rather than rebuilds, that accident stops
    # working. The text is passed through explicitly instead.
    edits = _edits_from_body(body)
    _file_edits(db, game, coach, edits, bool((body or {}).get("remember")),
                models.GameScoutingCorrection,
                remember_team=((body or {}).get("remember_team") or None))
    job = genjob.start(db, coach.id, "scouting",
                       {"game_id": game_id, "coach_id": coach.id, "edits": edits})
    background_tasks.add_task(_run_scouting_job, game_id, coach.id, job.id, edits)
    return {"job_id": job.id}


@router.post("/sessions/{game_id}/ai-scouting")
async def generate_ai_scouting(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Generate inline and wait. Kept for browser tabs loaded before the job
    endpoint above existed; new callers should use that one."""
    # Any coach who can access the game (owner or team staff) can generate their
    # OWN scouting report for it. Reports are per-coach and private until shared.
    game = _get_game_readable(db, game_id, coach)
    text = await _run_scouting(db, coach, game, [])
    return {"ai_scouting_report": text}


async def _run_game_report(db: Session, coach: models.Coach, game: models.GameSession,
                           corrections: list[str], on_words=None) -> str:
    """Build + run the full GAME REPORT (our team + opponent) from the box score
    plus opponent notes and any coach-added context, persist it for this coach,
    and return the text. Mirrors _run_scouting but covers both sides."""
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not configured.")

    def _side_context(is_opp: bool) -> str:
        summary: dict[str, dict] = {}
        for s in game.player_stats:
            if s.is_opponent != is_opp:
                continue
            d = summary.setdefault(s.player_name, {"off": 0.0, "def": 0.0, "stats": defaultdict(int)})
            d[("off" if s.stat_category == "offense" else "def")] += s.weighted_points
            d["stats"][s.stat_name] += s.count
        lines = ""
        for pname, d in summary.items():
            top = sorted(d["stats"].items(), key=lambda x: x[1], reverse=True)[:5]
            lines += f"\n{pname}: OFF={d['off']:.1f} DEF={d['def']:.1f}  {', '.join(f'{s}={c}' for s, c in top)}"
        return lines or "\n(no tracked stats)"

    team_row = db.get(models.Team, game.team_id) if game.team_id else None
    my_team = team_row.name if team_row else coach.program_name
    score_info = ""
    if game.our_score is not None and game.opponent_score is not None:
        res = "WIN" if game.our_score > game.opponent_score else ("LOSS" if game.our_score < game.opponent_score else "TIE")
        score_info = f"Final: {game.our_score}-{game.opponent_score} ({res})"

    context = (_team_notes_text(db, coach, game) + film_notes_for_game(db, coach, game)
               + team_material_block(db, coach, game.opponent_name or "")
               + learned_for_game(db, coach, game))
    if corrections:
        context += (
            "\n\nCOACH CONTEXT & ADJUSTMENTS (qualitative detail the box score can't capture — "
            "you MUST weave ALL of these into the analysis):\n"
            + "\n".join(f"- {c}" for c in corrections if c and c.strip())
        )

    from ..coach_context import resolve_level, language_directive
    _lvl = resolve_level(coach, team=team_row)
    prompt = (
        f"You are the BloomPrint Basketball Intelligence Model. Generate a full GAME REPORT for "
        f"{my_team} vs {game.opponent_name}.\n\nGame date: {game.date}\n"
        f"COMPETITION LEVEL: {_lvl} — calibrate every grade, comparison, and recommendation to this level.\n"
        f"{score_info}\n\n"
        f"OUR TEAM PLAYER GRADES:{_side_context(False)}\n\n"
        f"OPPONENT PLAYER GRADES:{_side_context(True)}"
        f"{context}\n\n"
        "Cover, in this order: 1) OUR TEAM PERFORMANCE — what worked, who stood out, where we broke down, "
        "and adjustments for next time; 2) OPPONENT BREAKDOWN — their tendencies, key players, how to attack "
        "and defend them going forward; 3) KEY TAKEAWAYS."
        f"{REPORT_FORMAT_WITH_TABLES}"
        f"{language_directive(coach)}"
    )
    try:
        # Streamed, for the reason given in _run_scouting.
        raw = await long_text(prompt, max_tokens=12000, on_words=on_words)
        if not raw.strip():
            raise HTTPException(status_code=500, detail="AI returned no content")
        import re as _re
        text = _re.sub(r"\s*END OF REPORT\.?\s*$", "", raw.strip(), flags=_re.IGNORECASE).rstrip()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {exc}")
    _upsert_game_report(db, coach, game, text)
    db.commit()
    return text


def _game_report_corrections(db: Session, game_id: int, coach_id: int) -> list[str]:
    rows = (
        db.query(models.GameSessionReportCorrection)
        .filter_by(game_id=game_id, coach_id=coach_id)
        .order_by(models.GameSessionReportCorrection.id)
        .all()
    )
    return [c.correction for c in rows]


def _run_game_report_job(game_id: int, coach_id: int, job_id: int,
                         extra_edits: list[str] | None = None) -> None:
    """The full game report, off the request that asked for it. See api/genjob.py."""
    import asyncio

    def work():
        db = SessionLocal()
        try:
            game = db.get(models.GameSession, game_id)
            coach = db.get(models.Coach, coach_id)
            if not game or not coach:
                raise RuntimeError("Game not found")
            rows = (db.query(models.GameSessionReportCorrection)
                      .filter_by(game_id=game_id, coach_id=coach_id)
                      .order_by(models.GameSessionReportCorrection.id).all())
            prior = _game_report_text_for(db, game_id, coach_id)
            _snapshot_report(db, game_id, coach_id, "game_report", prior)
            edits = list(extra_edits or []) + [c.correction for c in rows if not c.applied]
            if prior and edits:
                text = asyncio.run(_apply_edits(prior, edits, coach,
                                                on_words=genjob.words_reporter(job_id)))
                _upsert_game_report(db, coach, game, text.strip())
            else:
                asyncio.run(_run_game_report(db, coach, game, [c.correction for c in rows],
                                             on_words=genjob.words_reporter(job_id)))
            for r in rows:
                r.applied = True
            db.commit()
        finally:
            db.close()
        return game_id

    genjob.run(job_id, work)


@router.post("/sessions/{game_id}/game-report-job")
def start_full_game_report(
    game_id: int,
    background_tasks: BackgroundTasks,
    body: dict | None = None,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Start the full game report and hand back a job to follow.

    Serves both the first generation and "Apply & Regenerate"; see the scouting
    equivalent above.
    """
    game = _get_game_readable(db, game_id, coach)
    edits = _edits_from_body(body)
    _file_edits(db, game, coach, edits, bool((body or {}).get("remember")),
                models.GameSessionReportCorrection,
                remember_team=((body or {}).get("remember_team") or None))
    job = genjob.start(db, coach.id, "game_report_full",
                       {"game_id": game_id, "coach_id": coach.id, "edits": edits})
    background_tasks.add_task(_run_game_report_job, game_id, coach.id, job.id, edits)
    return {"job_id": job.id}


@router.post("/sessions/{game_id}/game-report")
async def generate_game_report(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """A full GAME REPORT — our team's performance AND the opponent, combined
    (vs the Scout Opponent report which is opponent-only). Persisted per-coach so
    it appears in Recents and feeds game packets.

    Generates inline and waits; kept for browser tabs loaded before the job
    endpoint above existed."""
    game = _get_game_readable(db, game_id, coach)
    all_corr = (
        db.query(models.GameSessionReportCorrection)
        .filter_by(game_id=game.id, coach_id=coach.id)
        .order_by(models.GameSessionReportCorrection.id)
        .all()
    )
    text = await _run_game_report(db, coach, game, [c.correction for c in all_corr])
    return {"report_text": text, "ai_game_report": text}


@router.get("/sessions/{game_id}/game-report-corrections")
def list_game_report_corrections(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game_readable(db, game_id, coach)
    rows = (
        db.query(models.GameSessionReportCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id)
        .order_by(models.GameSessionReportCorrection.id)
        .all()
    )
    return [{"id": r.id, "correction": r.correction, "applied": r.applied, "created_at": r.created_at} for r in rows]


@router.post("/sessions/{game_id}/game-report-corrections")
def add_game_report_correction(
    game_id: int,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game_readable(db, game_id, coach)
    text = (body.get("text") or body.get("correction") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Correction text required")
    row = models.GameSessionReportCorrection(game_id=game_id, coach_id=coach.id, correction=text)
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "correction": row.correction, "applied": row.applied}


@router.delete("/game-report-corrections/{correction_id}")
def delete_game_report_correction(
    correction_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    row = db.get(models.GameSessionReportCorrection, correction_id)
    if not row or row.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Correction not found")
    if row.applied:
        raise HTTPException(status_code=400, detail="Already applied — can't delete")
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.post("/sessions/{game_id}/apply-game-report-corrections")
async def apply_game_report_corrections(
    game_id: int,
    body: dict = None,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Regenerate the game report from the box score PLUS all of the coach's
    added context, then mark those corrections applied."""
    game = _get_game_readable(db, game_id, coach)
    if body:
        feedback = (body.get("feedback") or body.get("text") or "").strip()
        if feedback:
            db.add(models.GameSessionReportCorrection(game_id=game_id, coach_id=coach.id, correction=feedback))
            from .preferences import remember
            remember(db, coach.id, feedback, game.team_id)
            db.commit()
    pending = (
        db.query(models.GameSessionReportCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id, applied=False)
        .order_by(models.GameSessionReportCorrection.id)
        .all()
    )
    all_corr = (
        db.query(models.GameSessionReportCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id)
        .order_by(models.GameSessionReportCorrection.id)
        .all()
    )
    text = await _run_game_report(db, coach, game, [c.correction for c in all_corr])
    for p in pending:
        p.applied = True
    db.commit()
    return {"ai_game_report": text}


@router.get("/sessions/{game_id}/scouting-corrections")
def list_scouting_corrections(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game_readable(db, game_id, coach)
    rows = (
        db.query(models.GameScoutingCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id)
        .order_by(models.GameScoutingCorrection.id)
        .all()
    )
    return [{"id": r.id, "correction": r.correction, "applied": r.applied, "created_at": r.created_at} for r in rows]


@router.post("/sessions/{game_id}/scouting-corrections")
def add_scouting_correction(
    game_id: int,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game_readable(db, game_id, coach)
    text = (body.get("text") or body.get("correction") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Correction text required")
    row = models.GameScoutingCorrection(game_id=game_id, coach_id=coach.id, correction=text)
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "correction": row.correction, "applied": row.applied}


@router.patch("/scouting-corrections/{correction_id}")
def edit_scouting_correction(
    correction_id: int,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    row = db.get(models.GameScoutingCorrection, correction_id)
    if not row or row.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Correction not found")
    if row.applied:
        raise HTTPException(status_code=400, detail="Already applied — can't edit")
    text = (body.get("text") or body.get("correction") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Correction text required")
    row.correction = text
    db.commit()
    return {"id": row.id, "correction": row.correction, "applied": row.applied}


@router.delete("/scouting-corrections/{correction_id}")
def delete_scouting_correction(
    correction_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    row = db.get(models.GameScoutingCorrection, correction_id)
    if not row or row.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Correction not found")
    if row.applied:
        raise HTTPException(status_code=400, detail="Already applied — can't delete")
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.post("/sessions/{game_id}/apply-scouting-corrections")
async def apply_scouting_corrections(
    game_id: int,
    body: dict = None,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Regenerate the scouting report from the box score PLUS all of the coach's
    added context, then mark those corrections applied."""
    game = _get_game_readable(db, game_id, coach)
    if body:
        feedback = (body.get("feedback") or body.get("text") or "").strip()
        if feedback:
            db.add(models.GameScoutingCorrection(game_id=game_id, coach_id=coach.id, correction=feedback))
            # Remembered against the opponent, who a scouting report is about.
            from .preferences import remember
            remember(db, coach.id, feedback, _opponent_team_id_of(db, coach, game))
            db.commit()
    pending = (
        db.query(models.GameScoutingCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id, applied=False)
        .order_by(models.GameScoutingCorrection.id)
        .all()
    )
    # Include ALL of the coach's context so the report reflects everything they've added.
    all_corr = (
        db.query(models.GameScoutingCorrection)
        .filter_by(game_id=game_id, coach_id=coach.id)
        .order_by(models.GameScoutingCorrection.id)
        .all()
    )
    text = await _run_scouting(db, coach, game, [c.correction for c in all_corr])
    for p in pending:
        p.applied = True
    db.commit()
    return {"ai_scouting_report": text}


def _team_filter(db: Session, param: str | None):
    """A SQL condition matching games the chosen teams were in — EITHER side.

    Filtering on team_id alone asked "games my Senegal Lions side played", and a
    team the coach has only ever faced answered with nothing at all. Teams here
    are used both ways — your own sides and the opponents you keep records on —
    so picking one means the team, not the column it happened to sit in.
    """
    from sqlalchemy import or_ as _or
    picked = _selected_team_ids(param)
    if not picked:
        return None
    names = [t.name for t in db.query(models.Team).filter(models.Team.id.in_(picked)).all() if t.name]
    conds = [models.GameSession.team_id.in_(picked)]
    if names:
        # Opponents are free text, so matched by name and case-insensitively —
        # "Duke" and "duke" are one opponent, as the New Game picker enforces.
        conds.append(func.lower(models.GameSession.opponent_name).in_([n.lower() for n in names]))
    return _or(*conds)


def _norm_team(x) -> str:
    """A team name reduced to letters and digits, for matching what was typed."""
    return "".join(ch for ch in str(x or "").lower() if ch.isalnum())


def _selected_team_ids(param: str | None) -> list[int] | None:
    """The teams a screen is asking about, or None for "all of them".

    Team Grade showed every game on every team a coach could reach, whatever the
    team picker said — the endpoints had no way to be asked for less. A coach
    with two teams saw one pooled season and a leaderboard mixing both rosters,
    with nothing on the row to say who belonged where.
    """
    if not param:
        return None
    ids = [int(x) for x in str(param).split(",") if str(x).strip().lstrip("-").isdigit()]
    return ids or None


# ── Season Dashboard ──────────────────────────────────────────────────────────

@router.get("/season-dashboard")
def season_dashboard(
    phases: str | None = None,   # comma-separated, e.g. "playoff,tournament"
    season_year: str | None = None,
    team_ids: str | None = None,  # comma-separated; omitted means every team
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    # Include games on teams the coach is linked to as staff, so a team's season
    # shows in the Team Grade dashboard for all its staff, not just the owner.
    from sqlalchemy import or_
    # Named apart from the team_ids QUERY PARAMETER above. When both were
    # called team_ids the local won, the filter was handed a list of ints, and
    # it silently parsed to "no selection" — the picker appeared to work and
    # changed nothing.
    accessible_ids = _accessible_team_ids(db, coach)
    conds = [models.GameSession.coach_id == coach.id]
    if accessible_ids:
        conds.append(models.GameSession.team_id.in_(accessible_ids))
    q = db.query(models.GameSession).filter(or_(*conds)).filter(models.GameSession.status == "completed")
    picked = _team_filter(db, team_ids)
    if picked is not None:
        q = q.filter(picked)
    if phases:
        phase_list = [p.strip() for p in phases.split(",") if p.strip()]
        if phase_list:
            q = q.filter(models.GameSession.season_phase.in_(phase_list))
    if season_year:
        q = q.filter(models.GameSession.season_year == season_year)
    games = q.order_by(models.GameSession.date).all()

    wins = 0
    losses = 0
    team_grade_trend = []
    # Keyed by (team, player), not by player. Two teams with a Chris on each had
    # their games pooled into one leaderboard row — and with several teams shown
    # together, a row with no team on it is unreadable anyway.
    player_totals: dict[tuple[int | None, str], dict] = {}
    team_names = {t.id: t.name for t in db.query(models.Team).all()}

    # Which side of each game the selection is actually asking about.
    #
    # The picker matches games a team was in on EITHER side, so choosing a team
    # you have only ever played brought their games back — graded from your
    # bench, with your players on the leaderboard, under their name. A season
    # for that team has to be read from their side of the same games.
    #
    # With nothing picked this stays our side: "all teams" means the coach's own
    # season, and counting each game twice would make the record meaningless.
    picked_ids = _selected_team_ids(team_ids)
    if picked_ids:
        subject_ids = set(picked_ids)
        subject_names = {(t.name or "").strip().lower() for t in
                         db.query(models.Team).filter(models.Team.id.in_(picked_ids)).all()}
    else:
        # "All teams" means all of the COACH'S teams, not every team on file.
        # Once a team the coach only keeps records on can be either side of a
        # game, a scouted Egypt-vs-Senegal would otherwise walk into their own
        # win-loss record.
        mine = [t for t in db.query(models.Team).filter_by(coach_id=coach.id).all()
                if t.is_mine is not False]
        subject_ids = {t.id for t in mine}
        subject_names = {(t.name or "").strip().lower() for t in mine}

    # Minutes for every game at once. Asked per game, this was one round trip
    # per row of the trend chart — and most games have no minutes recorded at
    # all, so it was a query to learn nothing.
    minutes_by: dict[tuple[int, bool], dict[str, float]] = {}
    if games:
        for r in (db.query(models.GameMinutesPlayed)
                    .filter(models.GameMinutesPlayed.game_id.in_([g.id for g in games])).all()):
            minutes_by.setdefault((r.game_id, bool(r.is_opponent)), {})[r.player_name] = r.minutes_played
    # Stats too: game.player_stats is a lazy relationship, so touching it inside
    # the loop is another query per game.
    #
    # Columns, not entities. A season is thousands of stat rows and this needs
    # six fields off each — building full ORM objects for them costs more than
    # the queries it saves. A SQLAlchemy Row exposes the same attribute names,
    # so _compute_grades cannot tell the difference.
    stats_by: dict[int, list] = {}
    if games:
        cols = (models.GamePlayerStat.game_id, models.GamePlayerStat.player_name,
                models.GamePlayerStat.jersey_number, models.GamePlayerStat.is_opponent,
                models.GamePlayerStat.stat_name, models.GamePlayerStat.stat_category,
                models.GamePlayerStat.weighted_points, models.GamePlayerStat.quarter,
                models.GamePlayerStat.count)
        for st in (db.query(*cols)
                     .filter(models.GamePlayerStat.game_id.in_([g.id for g in games])).all()):
            stats_by.setdefault(st.game_id, []).append(st)

    for game in games:
        # Which side of this game the answer is about. A game whose own team is
        # not the subject, but whose opponent is, is read from the opponent's
        # bench — that is what makes a season for a team the coach only tracks.
        from_theirs = game.team_id not in subject_ids and \
            (game.opponent_name or "").strip().lower() in subject_names
        if game.team_id not in subject_ids and not from_theirs:
            continue    # neither side is what was asked about
        if game.our_score is not None and game.opponent_score is not None:
            our_pts, opp_pts = game.our_score, game.opponent_score
        else:
            our_pts, opp_pts = scores_from_stats(stats_by.get(game.id, []))
        if from_theirs:
            our_pts, opp_pts = opp_pts, our_pts
        if our_pts is not None and opp_pts is not None:
            if our_pts > opp_pts:
                wins += 1
            else:
                losses += 1
        win_loss_factor = _win_loss_factor(our_pts, opp_pts)

        our_stats = [s for s in stats_by.get(game.id, []) if bool(s.is_opponent) == from_theirs]
        our_minutes = minutes_by.get((game.id, from_theirs), {})
        player_grades = _compute_grades(our_stats, our_minutes)

        team_grade = _team_grade(player_grades, our_pts, opp_pts)
        # Whose season this row belongs to, and who they played — both flip when
        # the game is being read from the other bench.
        row_team_id = game.team_id
        row_team_name = team_names.get(game.team_id)
        row_opponent = game.opponent_name
        if from_theirs:
            row_team_name = game.opponent_name
            row_team_id = next((tid for tid, nm in team_names.items()
                                if (nm or "").strip().lower() == (game.opponent_name or "").strip().lower()),
                               None)
            row_opponent = team_names.get(game.team_id)

        team_grade_trend.append({
            "game_id": game.id,
            "opponent": row_opponent,
            "team_id": row_team_id,
            "team_name": row_team_name,
            "date": game.date.isoformat() if game.date else None,
            "team_grade": team_grade,
            "our_score": our_pts,
            "opponent_score": opp_pts,
        })

        for pg in player_grades:
            key = (row_team_id, pg["player_name"])
            if key not in player_totals:
                player_totals[key] = {"games": 0, "total_grade": 0.0, "total_off": 0.0, "total_def": 0.0}
            player_totals[key]["games"] += 1
            player_totals[key]["total_grade"] += pg["game_grade"]
            player_totals[key]["total_off"] += pg["offensive_grade"]
            player_totals[key]["total_def"] += pg["defensive_grade"]

    player_leaderboard = []
    for (tid, name), data in player_totals.items():
        g = max(data["games"], 1)
        player_leaderboard.append({
            "player_name": name,
            "team_id": tid,
            "team_name": team_names.get(tid),
            "avg_game_grade": round(data["total_grade"] / g, 2),
            "games_played": data["games"],
            "avg_offensive": round(data["total_off"] / g, 2),
            "avg_defensive": round(data["total_def"] / g, 2),
        })
    player_leaderboard.sort(key=lambda x: x["avg_game_grade"], reverse=True)

    total_games = wins + losses
    win_pct = round(wins / total_games, 3) if total_games > 0 else 0.0
    season_avg = round(
        sum(t["team_grade"] for t in team_grade_trend) / len(team_grade_trend), 2
    ) if team_grade_trend else 0.0

    # Phase breakdown
    all_games = db.query(models.GameSession).filter_by(coach_id=coach.id).all()
    phases = list({g.season_phase for g in all_games if g.season_phase})

    return {
        "record": {"wins": wins, "losses": losses, "win_pct": win_pct},
        "team_grade_trend": team_grade_trend,
        "player_leaderboard": player_leaderboard,
        "season_avg_team_grade": season_avg,
        "phases_available": phases,
    }


# ── Opponent Profile ──────────────────────────────────────────────────────────

SCOUT_SUBJECTS = {
    "offense": "what this team does on offense",
    "defense": "what this team does on defense",
    "weak": "where this team is vulnerable",
}


class ScoutInsightIn(BaseModel):
    # A player's name, or one of offense / defense / weak.
    subject: str
    refresh: bool = False


@router.post("/opponents/{opponent_name}/insight")
async def scout_insight(
    opponent_name: str,
    body: ScoutInsightIn,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """One sentence about a team, or about one of its players.

    Averages say what somebody did; a coach still has to decide what to do
    about it. This is that, in a sentence, from the numbers already on the
    page — and it is kept, so opening the same player twice costs one call,
    not two. It is rewritten when the game count behind it changes, because a
    line written from two games should not be presented beside four.

    A stored one is returned with `stale` when there is material it was not
    written from — a new packet report or film breakdown about the team. The
    page says so and offers to rewrite it, rather than either quietly showing
    an out-of-date line or spending a call per player behind the coach's back.
    """
    subject = (body.subject or "").strip()
    if not subject:
        raise HTTPException(status_code=400, detail="Which player or section?")

    profile = opponent_profile(opponent_name, db, coach)
    games_n = profile.get("games_count") or 0
    written = team_written_material(db, coach, opponent_name)
    material_n = len(written)

    row = (db.query(models.ScoutInsight)
             .filter_by(coach_id=coach.id, team_name=opponent_name, subject=subject)
             .order_by(models.ScoutInsight.id.desc()).first())
    if row and not body.refresh and row.games == games_n:
        return {"insight": row.insight, "games": row.games, "cached": True,
                # Kept, but flagged: there is written material about this team
                # that this sentence predates.
                "stale": (row.material or 0) != material_n,
                "material": row.material or 0, "material_now": material_n}

    def _lines(rows, key):
        return "; ".join(f"{r['stat']} {r[key]} per game" for r in rows) or "none recorded"

    if subject in SCOUT_SUBJECTS:
        section = {"offense": profile["offensive_tendencies"],
                   "defense": profile["defensive_tendencies"],
                   "weak": profile["weak_spots"]}[subject]
        facts = _lines(section, "per_game")
        ask = (f"{opponent_name} across {games_n} tracked game(s). "
               f"{SCOUT_SUBJECTS[subject].capitalize()}, by the numbers: {facts}.")
    else:
        who = next((p for p in profile["best_players"]
                    if p["player_name"].strip().lower() == subject.lower()), None)
        if not who:
            raise HTTPException(status_code=404, detail=f"No record of {subject} on this team.")
        a = who["averages"]
        facts = (f"{a['PTS']} pts, {a['REB']} reb, {a['AST']} ast, {a['STL']} stl, "
                 f"{a['BLK']} blk, {a['TO']} to per game; "
                 f"FG {a['FG_PCT'] if a['FG_PCT'] is not None else 'n/a'}%, "
                 f"3PT {a['THREE_PCT'] if a['THREE_PCT'] is not None else 'n/a'}%; "
                 f"grade {who['avg_grade']} out of 5")
        ask = (f"{subject} of {opponent_name}, across {who['games']} tracked game(s): {facts}.")

    # What has been written about them, not only what was counted. A packet
    # report says why a number happened; the number on its own does not.
    # Two pieces, trimmed. The output is ONE sentence: ten thousand characters
    # of reading buys nothing it can say and costs seconds the coach spends
    # watching a spinner. The newest material is the material that matters.
    reading = ""
    if written:
        parts = [f"\n[{kind}]\n{(text or '')[:1200]}" for kind, text in written[:2]]
        reading = ("\n\nWHAT HAS ALREADY BEEN WRITTEN ABOUT THIS TEAM:"
                   + "".join(parts))

    prompt = (
        "You are a basketball scout briefing a head coach before they play this team.\n\n"
        f"{ask}{reading}\n\n"
        "Write ONE sentence, under 200 characters, saying what this MEANS for the game plan "
        "— the thing a coach would say to their staff. Lead with the read, not the numbers; "
        "quote at most one figure, and only if it carries the point. No preamble, no "
        "markdown, no quotation marks. If the sample is one game, do not write as though it "
        "is a pattern."
    )
    try:
        import anthropic
        client = anthropic.AsyncAnthropic()
        # Opus, deliberately. This is one sentence, but it is the sentence a
        # coach reads off the scouting page and takes into a game plan — the
        # tier rule is about what the output is FOR, not how long it is. The
        # speed came from the prompt and from not blanking the page instead.
        resp = await client.messages.create(
            model=OPUS, max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        blocks = [b for b in resp.content if hasattr(b, "text")]
        text = (blocks[0].text if blocks else "").strip().strip('"')
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not write that insight: {exc}")
    if not text:
        raise HTTPException(status_code=500, detail="The insight came back empty.")

    if row:
        row.insight, row.games, row.created_at = text, games_n, datetime.utcnow()
        row.material = material_n
    else:
        db.add(models.ScoutInsight(coach_id=coach.id, team_name=opponent_name,
                                   subject=subject, insight=text, games=games_n,
                                   material=material_n))
    db.commit()
    return {"insight": text, "games": games_n, "cached": False,
            "stale": False, "material": material_n, "material_now": material_n}


@router.get("/opponents/{opponent_name}/insights")
def scout_insights(
    opponent_name: str,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    """Everything already written about this team, so the page can show what it
    has without paying for anything it does not need."""
    rows = (db.query(models.ScoutInsight)
              .filter_by(coach_id=coach.id, team_name=opponent_name)
              .order_by(models.ScoutInsight.id).all())
    # Counted once for the whole page rather than per row: it is the same
    # question for every sentence on it.
    material_n = team_material_count(db, coach, opponent_name)
    return {r.subject: {"insight": r.insight, "games": r.games,
                        "stale": (r.material or 0) != material_n,
                        "material": r.material or 0, "material_now": material_n}
            for r in rows}


@router.get("/opponents/{opponent_name}")
def opponent_profile(
    opponent_name: str,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    # Opponent intel is built from every game against them that the coach can
    # see — their own games plus games on teams they're staff on.
    from sqlalchemy import or_
    # Named apart from the team_ids QUERY PARAMETER above. When both were
    # called team_ids the local won, the filter was handed a list of ints, and
    # it silently parsed to "no selection" — the picker appeared to work and
    # changed nothing.
    accessible_ids = _accessible_team_ids(db, coach)
    conds = [models.GameSession.coach_id == coach.id]
    if accessible_ids:
        conds.append(models.GameSession.team_id.in_(accessible_ids))
    # Either side of the scoreboard. Matching only opponent_name meant a team
    # could have three games on file and no scouting page, because it happened
    # to be stored as the game's own team every time — which is a detail of how
    # the row is written, not something a coach knows or should have to.
    same = [tm.id for tm in db.query(models.Team).filter_by(coach_id=coach.id).all()
            if _norm_team(tm.name) == _norm_team(opponent_name)]
    side = [func.lower(models.GameSession.opponent_name) == opponent_name.strip().lower()]
    if same:
        side.append(models.GameSession.team_id.in_(same))
    games = (
        db.query(models.GameSession)
        .filter(or_(*conds), or_(*side))
        .order_by(models.GameSession.date.desc())
        .all()
    )
    if not games:
        raise HTTPException(status_code=404, detail="No games found for this team")

    player_totals: dict[str, dict] = {}
    offense_tendencies: dict[str, int] = defaultdict(int)
    defense_tendencies: dict[str, int] = defaultdict(int)
    latest_report = None

    # One pass for the stats and one for the minutes, rather than two per game.
    # See the season dashboard: columns rather than entities, because this reads
    # a handful of fields off thousands of rows.
    ids = [g.id for g in games]
    cols = (models.GamePlayerStat.game_id, models.GamePlayerStat.player_name,
            models.GamePlayerStat.jersey_number, models.GamePlayerStat.is_opponent,
            models.GamePlayerStat.stat_name, models.GamePlayerStat.stat_category,
            models.GamePlayerStat.weighted_points, models.GamePlayerStat.quarter,
            models.GamePlayerStat.count)
    stats_by: dict[int, list] = {}
    for st in db.query(*cols).filter(models.GamePlayerStat.game_id.in_(ids)).all():
        stats_by.setdefault(st.game_id, []).append(st)
    minutes_by: dict[tuple[int, bool], dict[str, float]] = {}
    for r in (db.query(models.GameMinutesPlayed)
                .filter(models.GameMinutesPlayed.game_id.in_(ids)).all()):
        minutes_by.setdefault((r.game_id, bool(r.is_opponent)), {})[r.player_name] = r.minutes_played
    scouting_by = {r.game_id: r for r in db.query(models.GameScoutingReport)
                   .filter(models.GameScoutingReport.game_id.in_(ids),
                           models.GameScoutingReport.coach_id == coach.id).all()}

    for game in games:
        # The coach's own scouting report for that game (not another coach's).
        row = scouting_by.get(game.id)
        own_scout = ((row.report_text if row and row.report_text else None)
                     or (game.ai_scouting_report if game.coach_id == coach.id else None))
        if own_scout:
            latest_report = own_scout
        # Which bench this team was on in THIS game. A team can be the game's
        # own team in one and the opponent in the next; reading is_opponent
        # blindly would scout whoever they happened to be playing.
        theirs = game.team_id not in same if same else True
        opp_stats = [s for s in stats_by.get(game.id, []) if bool(s.is_opponent) == theirs]
        for s in opp_stats:
            if s.player_name not in player_totals:
                player_totals[s.player_name] = {"games": 0, "counts": defaultdict(int)}
            player_totals[s.player_name]["counts"][s.stat_name] += (s.count or 0)
            if s.stat_category == "offense":
                offense_tendencies[s.stat_name] += s.count
            else:
                defense_tendencies[s.stat_name] += s.count
        # The 0-5 game grade, per player, the same way every other screen
        # computes it — so a number on this page means what it means elsewhere.
        for g in _compute_grades(opp_stats, minutes_by.get((game.id, theirs), {})):
            at = player_totals.setdefault(g["player_name"], {"games": 0, "counts": defaultdict(int)})
            at["games"] = at.get("games", 0) + 1
            at["grade_total"] = at.get("grade_total", 0.0) + g["game_grade"]
            at["jersey"] = at.get("jersey") or g.get("jersey_number")

    def _averages(counts: dict, games: int) -> dict:
        """Per game, from the same canonical line the box score prints."""
        line = _player_line(counts)
        n = max(games, 1)
        per = {k: round(line[k] / n, 1) for k in
               ("PTS", "REB", "AST", "STL", "BLK", "TO")}
        per["FG_PCT"] = round(100 * line["FGM"] / line["FGA"], 1) if line["FGA"] else None
        per["THREE_PCT"] = round(100 * line["3PM"] / line["3PA"], 1) if line["3PA"] else None
        return per

    # The scouted team's roster, so a number typed by hand shows here too.
    scout_roster: dict[str, str] = {}
    for tid in (same or []):
        scout_roster.update(_roster_jerseys(db, tid))

    best_players = sorted(
        [{"player_name": n,
          "jersey_number": _jersey_for(n, {n: d.get("jersey")} if d.get("jersey") else {},
                                       scout_roster),
          "avg_grade": round(d.get("grade_total", 0.0) / max(d["games"], 1), 2),
          "games": d["games"],
          "averages": _averages(d["counts"], d["games"])}
         for n, d in player_totals.items()],
        key=lambda x: x["avg_grade"],
        reverse=True,
    )[:5]

    n_games = max(len(games), 1)

    def _rate(pairs):
        """Counts as per-game numbers. A raw total says nothing without
        knowing how many games it took to reach."""
        return [{"stat": st, "count": ct, "per_game": round(ct / n_games, 1)}
                for st, ct in pairs]

    top_offense = sorted(offense_tendencies.items(), key=lambda x: x[1], reverse=True)[:3]
    top_defense = sorted(defense_tendencies.items(), key=lambda x: x[1], reverse=True)[:3]

    # Weak spots: the stats costing this team the most.
    #
    # Read from the same bench as everything else on the page. This block took
    # is_opponent at face value, so on a team that was the game's OWN team it
    # was quietly listing the other side's weaknesses under their name.
    all_stat_scores: dict[str, float] = defaultdict(float)
    all_stat_counts: dict[str, int] = defaultdict(int)
    for game in games:
        theirs = game.team_id not in same if same else True
        for s in stats_by.get(game.id, []):
            if bool(s.is_opponent) == theirs:
                all_stat_scores[s.stat_name] += s.weighted_points
                all_stat_counts[s.stat_name] += (s.count or 0)
    weak_spots = sorted(all_stat_scores.items(), key=lambda x: x[1])[:3]

    def _seen_from(g):
        ours, theirs_pts = effective_scores(g)
        # Printed from the READER's side, so a team's own scouting page never
        # shows their score in the other column. effective_scores() answers
        # from the game's own team's bench, which is this team only sometimes.
        return (ours, theirs_pts) if (same and g.team_id in same) else (theirs_pts, ours)

    team_names = {tm.id: tm.name for tm in db.query(models.Team).all()}
    games_list = []
    for g in games:
        a, b = _seen_from(g)
        # Who they played. A list of dates and scores with no names on it made a
        # scouting page unreadable the moment a team had more than one game.
        other = (g.opponent_name if (same and g.team_id in same)
                 else team_names.get(g.team_id) or g.opponent_name)
        games_list.append({"id": g.id, "date": g.date.isoformat() if g.date else None,
                           "opponent": other,
                           "our_score": a, "opponent_score": b, "status": g.status})

    return {
        "opponent_name": opponent_name,
        "games_played_against": games_list,
        "best_players": best_players,
        "games_count": len(games),
        "offensive_tendencies": _rate(top_offense),
        "defensive_tendencies": _rate(top_defense),
        "weak_spots": [{"stat": st, "score": round(sc, 2),
                        "per_game": round(all_stat_counts.get(st, 0) / n_games, 1)}
                       for st, sc in weak_spots],
        "ai_scouting_report": latest_report,
    }


# ── Opponent Notes ────────────────────────────────────────────────────────────

@router.get("/opponents/{opponent_name}/notes")
def get_opponent_notes(
    opponent_name: str,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    notes = (
        db.query(models.OpponentNote)
        .filter_by(coach_id=coach.id, opponent_name=opponent_name)
        .order_by(models.OpponentNote.created_at)
        .all()
    )
    return [{"id": n.id, "note_text": n.note_text, "created_at": n.created_at.isoformat()} for n in notes]


@router.post("/opponents/{opponent_name}/notes")
def add_opponent_note(
    opponent_name: str,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    note = models.OpponentNote(
        coach_id=coach.id,
        opponent_name=opponent_name,
        note_text=body.get("note_text", "").strip(),
    )
    if not note.note_text:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="note_text required")
    db.add(note)
    db.commit()
    db.refresh(note)
    return {"id": note.id, "note_text": note.note_text, "created_at": note.created_at.isoformat()}


@router.delete("/opponent-notes/{note_id}")
def delete_opponent_note(
    note_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    note = db.get(models.OpponentNote, note_id)
    if not note or note.coach_id != coach.id:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Note not found")
    db.delete(note)
    db.commit()
    return {"ok": True}


# ── Opponent Roster (saved opponent players, persisted by opponent name) ──────

def _opp_player_out(p: models.OpponentPlayer) -> dict:
    return {
        "id": p.id,
        "opponent_name": p.opponent_name,
        "player_name": p.player_name,
        "jersey_number": p.jersey_number,
        "position": p.position,
    }


@router.get("/opponents/{opponent_name}/players")
def list_opponent_players(
    opponent_name: str,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    players = (
        db.query(models.OpponentPlayer)
        .filter_by(coach_id=coach.id, opponent_name=opponent_name)
        .order_by(models.OpponentPlayer.id)
        .all()
    )
    return [_opp_player_out(p) for p in players]


@router.post("/opponents/{opponent_name}/players")
def add_opponent_player(
    opponent_name: str,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    name = (body.get("player_name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="player_name required")
    # Avoid duplicates for the same opponent
    existing = (
        db.query(models.OpponentPlayer)
        .filter_by(coach_id=coach.id, opponent_name=opponent_name, player_name=name)
        .first()
    )
    if existing:
        existing.jersey_number = (body.get("jersey_number") or "").strip() or existing.jersey_number
        existing.position = (body.get("position") or "").strip() or existing.position
        db.commit()
        db.refresh(existing)
        return _opp_player_out(existing)
    player = models.OpponentPlayer(
        coach_id=coach.id,
        opponent_name=opponent_name,
        player_name=name,
        jersey_number=(body.get("jersey_number") or "").strip() or None,
        position=(body.get("position") or "").strip() or None,
    )
    db.add(player)
    db.commit()
    db.refresh(player)
    return _opp_player_out(player)


@router.delete("/opponent-players/{player_id}")
def delete_opponent_player(
    player_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    p = db.get(models.OpponentPlayer, player_id)
    if not p or p.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Opponent player not found")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ── Compare Games ─────────────────────────────────────────────────────────────

@router.get("/compare")
def compare_games(
    game1_id: int,
    game2_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    game1 = _get_game(db, game1_id, coach.id)
    game2 = _get_game(db, game2_id, coach.id)

    def _summarize(game: models.GameSession) -> dict:
        our_stats = [s for s in game.player_stats if not s.is_opponent]
        mp_records = db.query(models.GameMinutesPlayed).filter_by(game_id=game.id, is_opponent=False).all()
        minutes = {r.player_name: r.minutes_played for r in mp_records}
        grades = _compute_grades(our_stats, minutes)
        avg_pg = sum(g["game_grade"] for g in grades) / len(grades) if grades else 0.0
        if game.our_score is not None and game.opponent_score is not None:
            wlf = 10.0 if game.our_score > game.opponent_score else (7.0 if abs(game.our_score - game.opponent_score) <= 5 else 5.0)
        else:
            wlf = 5.0
        team_grade = round(avg_pg * 0.6 + wlf * 0.4, 2)
        return {
            "game_id": game.id,
            "opponent": game.opponent_name,
            "date": game.date.isoformat() if game.date else None,
            "our_score": game.our_score,
            "opponent_score": game.opponent_score,
            "team_grade": team_grade,
            "player_grades": grades,
        }

    return {
        "game1": _summarize(game1),
        "game2": _summarize(game2),
    }


# ── Whiteboards ───────────────────────────────────────────────────────────────

@router.get("/sessions/{game_id}/whiteboards")
def list_whiteboards(
    game_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game(db, game_id, coach.id)
    boards = db.query(models.GameWhiteboard).filter_by(game_id=game_id, coach_id=coach.id).order_by(models.GameWhiteboard.created_at).all()
    return [{"id": b.id, "name": b.name, "court_type": b.court_type, "data": b.data, "created_at": b.created_at.isoformat()} for b in boards]


# Coach-level PLAYBOOK boards — not tied to any game, so plays the coach draws
# persist in "my boards" until deleted, regardless of games.
#
# NULL, not a 0 sentinel. game_id is a foreign key to game_sessions and no
# session has id 0: SQLite does not enforce foreign keys by default so this
# saved locally, while the deployed Postgres rejected every playbook write with
# a constraint violation surfacing as a 500. Rows written under the old sentinel
# are still read below so nothing drawn before this is lost.
_PLAYBOOK_GID = None
_LEGACY_PLAYBOOK_GID = 0


def _playbook_filter():
    """Match a coach's playbook boards, written under either convention."""
    from sqlalchemy import or_
    return or_(models.GameWhiteboard.game_id.is_(None),
               models.GameWhiteboard.game_id == _LEGACY_PLAYBOOK_GID)


@router.get("/playbook/whiteboards")
def list_playbook(
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    boards = (
        db.query(models.GameWhiteboard)
        .filter(_playbook_filter(), models.GameWhiteboard.coach_id == coach.id)
        .order_by(models.GameWhiteboard.created_at)
        .all()
    )
    return [{"id": b.id, "name": b.name, "court_type": b.court_type, "data": b.data, "created_at": b.created_at.isoformat()} for b in boards]


@router.post("/playbook/whiteboards")
def create_playbook(
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    board = models.GameWhiteboard(
        game_id=_PLAYBOOK_GID,
        coach_id=coach.id,
        name=body.get("name", "Untitled Board"),
        court_type=body.get("court_type", "full"),
        data=body.get("data", "[]"),
    )
    db.add(board)
    db.commit()
    db.refresh(board)
    return {"id": board.id, "name": board.name, "court_type": board.court_type, "data": board.data, "created_at": board.created_at.isoformat()}


@router.post("/sessions/{game_id}/whiteboards")
def create_whiteboard(
    game_id: int,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    _get_game(db, game_id, coach.id)
    board = models.GameWhiteboard(
        game_id=game_id,
        coach_id=coach.id,
        name=body.get("name", "Untitled Board"),
        court_type=body.get("court_type", "full"),
        data=body.get("data", "[]"),
    )
    db.add(board)
    db.commit()
    db.refresh(board)
    return {"id": board.id, "name": board.name, "court_type": board.court_type, "data": board.data, "created_at": board.created_at.isoformat()}


@router.patch("/whiteboards/{board_id}")
def update_whiteboard(
    board_id: int,
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    board = db.get(models.GameWhiteboard, board_id)
    if not board or board.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Whiteboard not found")
    if "name" in body:
        board.name = body["name"]
    if "court_type" in body:
        board.court_type = body["court_type"]
    if "data" in body:
        board.data = body["data"]
    board.updated_at = datetime.utcnow()
    db.commit()
    return {"id": board.id, "name": board.name, "court_type": board.court_type, "data": board.data}


# ── AI play validation (shared by draw-up / describe / adapt) ────────────────
def _pt(v, lo, hi, d):
    try:
        return max(lo, min(hi, float(v)))
    except (TypeError, ValueError):
        return d


def _clean_scheme(sc):
    sc = sc if isinstance(sc, dict) else {}
    out = {"players": [], "defenders": [], "actions": []}
    for i, pl in enumerate((sc.get("players") or [])[:5]):
        out["players"].append({"id": str(pl.get("id") or f"O{i+1}")[:3],
                               "x": _pt(pl.get("x"), -2, 52, 25), "y": _pt(pl.get("y"), 48, 96, 80),
                               "role": str(pl.get("role") or "")[:200]})
    for i, df in enumerate((sc.get("defenders") or [])[:5]):
        out["defenders"].append({"id": str(df.get("id") or f"X{i+1}")[:3],
                                 "x": _pt(df.get("x"), -2, 52, 25), "y": _pt(df.get("y"), 48, 96, 74),
                                 "role": str(df.get("role") or "")[:200]})

    # Auto-correct obviously-wrong on-ball defenders: a PERIMETER defender
    # (guarding a man up top, man.y < 76) that sits ON TOP of / above its man
    # (farther from the basket) is almost never right — nudge it to the basket
    # side (between man and rim). Post fronts / off-ball are left untouched.
    by_num = {p["id"][1:]: p for p in out["players"] if len(p["id"]) > 1}
    for df in out["defenders"]:
        man = by_num.get(df["id"][1:]) if len(df["id"]) > 1 else None
        if not man:
            continue
        perimeter = man["y"] < 76
        guarding = abs(df["x"] - man["x"]) < 10
        non_basket_side = df["y"] <= man["y"] + 1  # at/above man = away from rim
        if perimeter and guarding and non_basket_side:
            df["y"] = min(man["y"] + 5.0, 90.0)
            df["x"] = man["x"] + (25.0 - man["x"]) * 0.12  # ease slightly toward the middle
    for idx, a in enumerate((sc.get("actions") or [])[:10]):
        fr, to = a.get("from") or [25, 80], a.get("to") or [25, 70]
        try:
            step = max(1, int(a.get("step") or (idx + 1)))
        except (TypeError, ValueError):
            step = idx + 1
        out["actions"].append({"kind": str(a.get("kind") or "cut"), "step": step,
                               "actor": str(a.get("actor") or "")[:3],
                               "from": [_pt(fr[0], -2, 52, 25), _pt(fr[1], 48, 96, 80)],
                               "to":   [_pt(to[0], -2, 52, 25), _pt(to[1], 48, 96, 70)]})
    return out


def _clean_key(raw_key):
    out = []
    for i, k in enumerate((raw_key or [])[:5]):
        fr, to = (k.get("from") or [25, 80]), (k.get("to") or [25, 70])
        out.append({"n": int(k.get("n") or i + 1),
                    "text": str(k.get("text") or "")[:120],
                    "from": [_pt(fr[0], -2, 52, 25), _pt(fr[1], 48, 96, 80)],
                    "to":   [_pt(to[0], -2, 52, 25), _pt(to[1], 48, 96, 70)]})
    return out


# ── AI play draw-up ───────────────────────────────────────────────────────────
# Turns a coach's description of a scene/play into a structured X's-and-O's
# diagram: offense / defense / counter schemes plus a numbered improvement key.
_AI_PLAY_PROMPT = """You are an elite basketball tactician. From the scene description below, produce a schematic play diagram as STRICT JSON (no markdown, no prose).

POSITIONS: the player numbers are basketball positions — 1 = Point Guard (O1), 2 = Shooting Guard (O2), 3 = Small Forward (O3), 4 = Power Forward (O4), 5 = Center (O5). Place and move each player consistent with their position (e.g. O1 initiates up top, O5 plays around the rim/high post), and each defender X1-X5 guards the matching position (X5 guards O5, etc.).

COORDINATES: feet on a regulation HALF court. x: 0 = left sideline, 25 = middle, 50 = right sideline. y grows toward the hoop — LOW y is farther from the basket (up top), HIGH y is at the rim. The x,y for every player and defender is their STARTING position, before the play develops (movement is expressed only through actions).

DEFENDER POSITIONING: by default place each defender on the BASKET side of the player they guard — between their man and the rim. Because the rim is at HIGH y, that means the defender sits at a slightly HIGHER y than their man (about 3-5 ft toward the basket), roughly on the line from their man to the rim. So the on-ball defender of a point guard up top (y≈62) belongs just below him toward the basket (y≈66-67), NOT above him. Only put a defender on the non-basket (ball/deny) side when the scene or scheme calls for fronting the post, denying a wing, or top-locking a shooter. Landmarks: half-court line y=47, top of the 3-point arc / where the point guard initiates y≈62, free-throw line y=75, elbows y≈75 (x≈17 and x≈33), rim/basket y≈89, blocks/low post y≈84 (x≈19 and x≈31), wings y≈66 (x≈8 and x≈42), corners y≈90 (x≈4 and x≈46). The ball-handler up top belongs around y=60-64, NOT at the free-throw line. Keep in-bounds coordinates inside 2..48 for x and 48..92 for y.

INBOUND PLAYS (BLOB / SLOB): the court now shows a little out-of-bounds floor beyond the lines. If the scene is an inbound play, place the player taking the ball out (the inbounder) OUT OF BOUNDS and start the inbound pass from there: for a baseline out-of-bounds (BLOB) inbounder use y≈95 (just past the baseline) at the x you want along the endline; for a sideline out-of-bounds (SLOB) inbounder use x≈-1 (left) or x≈51 (right) at the y you want. Only use these out-of-bounds spots for an actual inbounder on an inbound play; everyone else stays in bounds.

Return exactly this shape. Output play_name and the FULL key array FIRST, then the schemes last — the key must never be omitted:
{
  "play_name": "short name",
  "key": [{"n": 1, "text": "concise coaching suggestion", "from": [x, y], "to": [x, y]}],
  "schemes": {
    "offense": {"players": [{"id": "O1", "x": 25, "y": 80, "role": "what O1 does in this scheme"}], "defenders": [{"id": "X1", "x": 25, "y": 76, "role": "what X1 does in this scheme"}], "actions": [{"actor": "O2", "kind": "cut|screen|pass|dribble", "from": [25, 80], "to": [30, 70], "step": 1}]},
    "defense": { same shape },
    "counter": { same shape }
  }
}

ACTOR: every action has an "actor" = the id of the player or defender who makes that movement (must match one of the players/defenders ids, e.g. "O2" or "X1"). The action's "from" should be that actor's current position and "to" is where they move. For a "pass", the actor is the passer and "from"/"to" are the ball's path (no one relocates on a pass).

STEPS: every action has a "step" integer (1,2,3,...) marking WHEN it happens as the play develops. Actions that happen at the SAME time share the same step number; actions that happen later get higher step numbers. Order the play realistically (e.g. screen on step 1, cut on step 2, pass on step 3).

ROLE: for EVERY player (O1-O5) AND defender (X1-X5) in EVERY scheme, include a short "role" string (under 120 characters) describing what that player does in that scheme — their action, assignment, or spot. Reflect the scene description and any BY-PLAYER INSTRUCTIONS so the coach can see how you read their words for each player.

BY-PLAYER INSTRUCTIONS: the scene may list per-player intent as O1-O5 (offense) and D1-D5 (defense). D1-D5 are the defenders X1-X5 (D1 = X1 guards O1, D2 = X2 guards O2, and so on). Assign each player/defender to do exactly what its note says, and reflect it in that player's role and movements.

STANDING PER-PLAYER INSTRUCTIONS: if the scene lists these, ALWAYS honor them — you may reword them for clarity but never drop the coach's intent, and reflect each one in that player's role and movements.

LABELS: in the human-facing "text" of every key item, refer to defenders as D1-D5, never X1-X5 (the JSON ids stay X1-X5, but the key text the coach reads must say D1-D5).

CONSTRAINTS: if the scene lists LOCKED PLAYERS or PER-PLAYER GUIDANCE / PLAYER ADJUSTMENTS, honor them precisely — keep any locked player at the exact coordinates given and reposition (cascade) only the other players to make the play work, and follow each per-player note.

SCHEME CONSISTENCY (important): the three schemes are LAYERS of ONE play, not three separate plays. By default OFFENSE is the base:
- OFFENSE = the base offensive play: the offensive players (O1-O5) and their movement.
- DEFENSE = the EXACT same offensive players, at the EXACT same starting positions, with the EXACT same offensive actions as OFFENSE (identical — do not change them), PLUS the defenders (X1-X5) and their reactions/rotations. (So the coach sees the offense the defense is reacting to.)
- COUNTER = the SAME offensive starting positions and the SAME early actions as OFFENSE, but it DIVERGES near the end into a different finish/read (the alternate outcome). Keep the beginning identical; only the last step(s) differ.
If the scene is a DEFENSE-FIRST request (the coach is drawing up a defense, e.g. "show a 2-3 zone", "how do we guard this"), then DEFENSE is the base instead, and OFFENSE/COUNTER layer onto that same defensive alignment.

DEFENSIVE MOVEMENT (defense scheme) — the defenders must ACTIVELY MOVE to stop a basket, never stand still. Give the BEST realistic defense against the offense's actions: the on-ball defender slides and mirrors the ball-handler; every off-ball defender MOVES WITH their man and stays between their man and the rim; on a ball screen pick a coverage (hedge, switch, drop, or blitz) and move the defenders to execute it; help off the weakest shooter to wall up drives and TAG the roller, then recover/close out; deny the primary scoring option; contest the shot at the rim. Add a stepped defensive action for EVERY defender who must move to prevent the score, timed (same step) to the offensive action they are reacting to.

RULES: 5 offensive players (O1-O5) and up to 5 defenders (X1-X5) per scheme; at most 10 actions per scheme; 3-5 key items. The key holds the suggested movements or positioning that would have made the play succeed (get a basket, find the open man, correct the rotation) — each key arrow shows the improved movement. Keep key text under 90 characters.

OUTPUT ONLY the JSON object. It MUST be strictly valid: every array element and object property separated by a comma, no trailing commas, no comments, no prose before or after."""


def _parse_ai_play_json(raw: str):
    """Best-effort parse of the model's JSON, repairing common LLM mistakes."""
    import re as _re
    import json as _json
    raw = _re.sub(r"^```(?:json)?|```$", "", raw, flags=_re.MULTILINE).strip()
    m = _re.search(r"\{.*\}", raw, _re.DOTALL)
    candidate = m.group(0) if m else raw
    try:
        return _json.loads(candidate)
    except Exception:
        pass
    # Repair: strip trailing commas before } or ]
    repaired = _re.sub(r",\s*([}\]])", r"\1", candidate)
    # Repair: insert missing commas between adjacent tokens across a newline.
    repaired = _re.sub(r"([}\]0-9\"])\s*\n\s*([{\[\"])", r"\1,\n\2", repaired)
    # Repair: missing comma between a value and the next object/array on the SAME
    # line (e.g. "} {" or "] [" or "5 {") — common when the model omits a comma.
    repaired = _re.sub(r"([}\]0-9])\s+([{\[])", r"\1,\2", repaired)
    return _json.loads(repaired)


@router.post("/ai-play")
async def ai_play(
    body: dict,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY not configured")
    description = (body.get("description") or "").strip()
    if not description:
        raise HTTPException(status_code=400, detail="Describe the scene or play first")
    # Fold any newly-drawn boards into the coach's style profile, then bias the
    # generation toward how this coach positions players.
    style_block = ""
    try:
        from ..play_style import refresh_and_get_profile
        profile = refresh_and_get_profile(db, coach)
        if profile and profile.strip():
            style_block = (
                "\n\nCOACH STYLE — how THIS coach positions players and what they favor. Bias the play "
                "toward this (especially when the scene is brief), unless the scene says otherwise:\n"
                f"{profile.strip()}\n"
            )
    except Exception:
        style_block = ""
    try:
        import anthropic
        client = anthropic.AsyncAnthropic()
        resp = await client.messages.create(
            model=OPUS,
            max_tokens=8000,
            messages=[{"role": "user", "content": f"{_AI_PLAY_PROMPT}{style_block}\n\nSCENE:\n{description}"}],
        )
        blocks = [b for b in resp.content if hasattr(b, "text")]
        raw = (blocks[0].text if blocks else "{}").strip()
        try:
            data = _parse_ai_play_json(raw)
        except Exception:
            # One retry: ask the model to return corrected, strictly-valid JSON.
            fix = await client.messages.create(
                model=OPUS,
                max_tokens=8000,
                messages=[{"role": "user", "content":
                    "The following was supposed to be strict JSON but is invalid. "
                    "Return ONLY the corrected, strictly-valid JSON — same data, no prose, "
                    f"no trailing commas, no comments:\n\n{raw}"}],
            )
            fblocks = [b for b in fix.content if hasattr(b, "text")]
            data = _parse_ai_play_json(fblocks[0].text if fblocks else "{}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AI play generation failed: {exc}")

    # Validate + clamp everything the client will render (helpers at module level).
    schemes = data.get("schemes") if isinstance(data.get("schemes"), dict) else {}
    key_items = _clean_key(data.get("key"))

    # Guarantee the key: if it came back empty, generate it in a separate,
    # small call so a long play can never leave the suggestions blank.
    if not key_items:
        try:
            import json as _json2
            kresp = await client.messages.create(
                model=OPUS,
                max_tokens=2000,
                messages=[{"role": "user", "content":
                    "For the basketball play below, list 3-5 KEY suggested improvements — the movements or "
                    "positioning that would have made it succeed (get a basket, find the open man, correct the "
                    "rotation). Return ONLY strict JSON: {\"key\": [{\"n\":1, \"text\":\"...\", \"from\":[x,y], "
                    "\"to\":[x,y]}]}. Coordinates are half-court feet (x 2..48, y 48..92, high y = rim). Key text "
                    f"under 90 chars.\n\nPLAY:\n{description}"}],
            )
            kb = [b for b in kresp.content if hasattr(b, "text")]
            kdata = _parse_ai_play_json(kb[0].text if kb else "{}")
            key_items = _clean_key(kdata.get("key"))
        except Exception:
            pass

    return {
        "play_name": str(data.get("play_name") or "AI Play")[:40],
        "schemes": {name: _clean_scheme(schemes.get(name)) for name in ("offense", "defense", "counter")},
        "key": key_items,
    }


# ── Re-describe after a manual player drag ────────────────────────────────────
# The coach dragged one player to a new spot. Keep everyone else exactly where
# they are; rewrite just the moved player's role and refresh the key.
@router.post("/ai-play-describe")
async def ai_play_describe(
    body: dict,
    coach: models.Coach = Depends(get_current_coach),
):
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY not configured")
    schemes = body.get("schemes") if isinstance(body.get("schemes"), dict) else {}
    scheme = str(body.get("scheme") or "offense")
    pid = str(body.get("player_id") or "")
    source = str(body.get("source") or "")[:2000]
    scm = schemes.get(scheme) or {}
    units = list(scm.get("players") or []) + list(scm.get("defenders") or [])
    moved = next((u for u in units if str(u.get("id")) == pid), None)
    if not moved:
        raise HTTPException(status_code=400, detail="Moved player not found in that scheme")

    def _disp(i):
        i = str(i or "")
        return ("D" + i[1:]) if i[:1] == "X" else i

    disp = _disp(pid)

    def _fmt(s):
        us = list(s.get("players") or []) + list(s.get("defenders") or [])
        return "\n".join(
            f"- {_disp(u.get('id'))} at x={round(float(u.get('x', 25)))}, y={round(float(u.get('y', 80)))}: {str(u.get('role') or '')[:120]}"
            for u in us
        )

    prompt = (
        "You are an elite basketball tactician. COORDINATES: half-court feet, x 0=left sideline / 50=right, "
        "y grows toward the rim (low y = up top, high y = at the basket; y>94 is out of bounds behind the baseline).\n\n"
        f"ORIGINAL PLAY: {source or '(none given)'}\n\n"
        f"CURRENT {scheme.upper()} FORMATION (everyone stays here — do NOT move anyone):\n{_fmt(scm)}\n\n"
        f"The coach just MANUALLY MOVED {disp} to x={round(float(moved.get('x', 25)))}, y={round(float(moved.get('y', 80)))}. "
        f"Rewrite ONLY what {disp} does from this new spot, consistent with the rest of the formation. "
        "Return STRICT JSON only, no prose: {\"role\": \"<one concise sentence under 120 chars>\", "
        "\"key\": [{\"n\": 1, \"text\": \"<refreshed coaching point under 90 chars, refer to defenders as D1-D5>\"}]}. "
        "Give 3-4 key items."
    )
    try:
        import anthropic
        client = anthropic.AsyncAnthropic()
        resp = await client.messages.create(
            model=OPUS, max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
        blocks = [b for b in resp.content if hasattr(b, "text")]
        data = _parse_ai_play_json(blocks[0].text if blocks else "{}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Re-describe failed: {exc}")

    role = str(data.get("role") or moved.get("role") or "")[:200]
    key = []
    for i, k in enumerate((data.get("key") or [])[:5]):
        key.append({"n": int(k.get("n") or i + 1), "text": str(k.get("text") or "")[:120]})
    return {"role": role, "key": key}


# ── Adapt a play around the coach's manual moves ──────────────────────────────
# The coach edited ONE scheme (moved players/arrows). Keep that scheme's starting
# positions locked and its movement minimally changed; then cascade the change to
# the DEPENDENT schemes (defense reacts to offense, counter adjusts to defense)
# and refresh the key, so the whole play stays consistent.
@router.post("/ai-play-adapt")
async def ai_play_adapt(
    body: dict,
    coach: models.Coach = Depends(get_current_coach),
):
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY not configured")
    # Back-compat: accept old single-scheme shape too.
    edited = str(body.get("edited") or body.get("scheme_name") or "offense")
    schemes = body.get("schemes") if isinstance(body.get("schemes"), dict) else {}
    if not schemes and isinstance(body.get("scheme"), dict):
        schemes = {edited: body["scheme"]}
    downstream = [s for s in (body.get("downstream") or []) if s in ("offense", "defense", "counter") and s != edited]
    key = body.get("key") if isinstance(body.get("key"), list) else []
    locked = body.get("locked") if isinstance(body.get("locked"), dict) else {}
    source = str(body.get("source") or "")[:2000]

    def _disp(i):
        i = str(i or "")
        return ("D" + i[1:]) if i[:1] == "X" else i

    def _pt2(pt):
        try:
            return f"({round(float(pt[0]))},{round(float(pt[1]))})"
        except Exception:
            return "(?,?)"

    def _fmt_scheme(name):
        sc = schemes.get(name) or {}
        units = "\n".join(
            f"    - {_disp(u.get('id'))} at ({round(float(u.get('x', 25)))},{round(float(u.get('y', 80)))}): {str(u.get('role') or '')[:90]}"
            for u in (list(sc.get("players") or []) + list(sc.get("defenders") or []))
        )
        acts = "\n".join(
            f"    - {_disp(a.get('actor'))} {a.get('kind')} {_pt2(a.get('from') or [0, 0])}->{_pt2(a.get('to') or [0, 0])} (step {a.get('step', 1)})"
            for a in (sc.get("actions") or [])
        )
        return f"  {name.upper()}:\n   players/defenders:\n{units or '    (none)'}\n   actions:\n{acts or '    (none)'}"

    all_txt = "\n\n".join(_fmt_scheme(n) for n in ("offense", "defense", "counter") if schemes.get(n))
    locked_units = ", ".join(f"{_disp(u.get('id'))}@{_pt2([u.get('x'), u.get('y')])}" for u in (locked.get("units") or [])) or "none"
    locked_actions = "; ".join(
        f"{_disp(a.get('actor'))} {a.get('kind')} {_pt2(a.get('from') or [0, 0])}->{_pt2(a.get('to') or [0, 0])}"
        for a in (locked.get("actions") or [])
    ) or "none"
    locked_keys = "; ".join(_pt2(k.get("from") or [0, 0]) + "->" + _pt2(k.get("to") or [0, 0]) for k in (locked.get("keys") or [])) or "none"
    dep_list = ", ".join(s.upper() for s in downstream) or "none"
    return_names = [edited] + downstream

    prompt = (
        "You are an elite basketball tactician. A play has three schemes — OFFENSE (what the offense runs), "
        "DEFENSE (the defensive scheme guarding it), COUNTER (the adjustment) — plus a KEY of suggested "
        "improvements. COORDINATES: half-court feet, x 0=left/50=right, y grows toward the rim (low y = up "
        "top, high y = at the basket; y>94 = out of bounds behind the baseline for inbounds).\n\n"
        f"ORIGINAL PLAY: {source or '(none given)'}\n\n"
        f"THE FULL PLAY NOW:\n{all_txt}\n\n"
        f"The coach just edited the {edited.upper()} scheme. Pieces the coach placed there (keep EXACTLY):\n"
        f"  players/defenders: {locked_units}\n  action arrows: {locked_actions}\n  key arrows: {locked_keys}\n\n"
        f"Do this:\n"
        f"1) EDITED scheme ({edited.upper()}) — MAKE THE SMALLEST CHANGE: keep EVERY player/defender at its "
        f"EXACT position (never move a starting spot), keep the existing action arrows, only adjust an arrow "
        f"if the coach's change broke it, and update role text.\n"
        f"2) DEPENDENT schemes ({dep_list}) — UPDATE these so they stay consistent with the edited "
        f"{edited.upper()}. If DEFENSE is dependent: keep the EXACT same offensive players, positions, and "
        f"offensive actions as the edited offense (identical), then give the defenders the BEST realistic "
        f"reaction to STOP a basket — the on-ball defender mirrors the ball, off-ball defenders move WITH "
        f"their man between man and rim, help/tag the roller on drives and screens, deny the primary "
        f"option, contest the shot — with a stepped defensive action for every defender who must move. If "
        f"COUNTER is dependent: keep the same offensive start and early actions, diverge only at the end. "
        f"Reposition/redraw within those rules.\n"
        f"3) Refresh the KEY to match the updated play.\n"
        f"Keep the same player/defender ids everywhere.\n\n"
        "Return STRICT JSON only, no prose. Include ONLY these schemes: "
        + ", ".join(f'"{n}"' for n in return_names) + ".\n"
        '{"schemes": {"' + edited + '": {"players": [{"id":"O1","x":25,"y":62,"role":"..."}], '
        '"defenders": [{"id":"X1","x":25,"y":66,"role":"..."}], '
        '"actions": [{"actor":"O2","kind":"cut|screen|pass|dribble","from":[x,y],"to":[x,y],"step":1}]}}, '
        '"key": [{"n":1,"text":"<under 90 chars, refer to defenders as D1-D5>","from":[x,y],"to":[x,y]}]}. '
        "Give 3-5 key items. Defender ids stay X1-X5 in JSON."
    )
    try:
        import anthropic
        client = anthropic.AsyncAnthropic()
        resp = await client.messages.create(
            model=OPUS, max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        blocks = [b for b in resp.content if hasattr(b, "text")]
        raw = (blocks[0].text if blocks else "{}").strip()
        try:
            data = _parse_ai_play_json(raw)
        except Exception:
            # One retry: ask the model to return corrected, strictly-valid JSON.
            fix = await client.messages.create(
                model=OPUS, max_tokens=4000,
                messages=[{"role": "user", "content":
                    "The following was supposed to be strict JSON but is invalid. Return ONLY the "
                    "corrected, strictly-valid JSON — same data, no prose, no trailing commas, no "
                    f"comments, every element comma-separated:\n\n{raw}"}],
            )
            fblocks = [b for b in fix.content if hasattr(b, "text")]
            data = _parse_ai_play_json(fblocks[0].text if fblocks else "{}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Adapt failed: {exc}")

    out_schemes = {}
    raw_schemes = data.get("schemes") if isinstance(data.get("schemes"), dict) else {}
    # Back-compat: a bare "scheme" applies to the edited scheme.
    if not raw_schemes and isinstance(data.get("scheme"), dict):
        raw_schemes = {edited: data["scheme"]}
    for name in return_names:
        if isinstance(raw_schemes.get(name), dict):
            out_schemes[name] = _clean_scheme(raw_schemes[name])
    out_key = _clean_key(data.get("key"))
    return {"schemes": out_schemes, "key": out_key}


# ── Review + name a hand-drawn (freehand) play ────────────────────────────────
# The coach played back a freehand drawing (players + ordered move/pass arrows).
# Return a short play NAME and a 1-2 sentence read of what happens, in order.
@router.post("/ai-play-name")
async def ai_play_name(
    body: dict,
    coach: models.Coach = Depends(get_current_coach),
):
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY not configured")
    markers = body.get("markers") if isinstance(body.get("markers"), list) else []
    arrows = body.get("arrows") if isinstance(body.get("arrows"), list) else []
    labels = [str(x)[:40] for x in (body.get("labels") or []) if x][:12]

    def _pt2(pt):
        try:
            return f"({round(float(pt[0]))},{round(float(pt[1]))})"
        except Exception:
            return "(?,?)"

    mk_txt = ", ".join(f"{('O' if m.get('kind') == 'O' else 'X')}@{_pt2([m.get('x'), m.get('y')])}" for m in markers[:12]) or "none marked"
    ar_txt = "\n".join(
        f"  {i+1}. step {a.get('step', i+1)}: {a.get('kind', 'move')} {_pt2(a.get('from') or [0, 0])} -> {_pt2(a.get('to') or [0, 0])}"
        for i, a in enumerate(sorted(arrows, key=lambda a: a.get('step', 0))[:20])
    ) or "  (no movement)"
    prompt = (
        "A basketball coach hand-drew a play on a half court (feet: x 0=left/50=right, y grows toward the "
        "rim). Players/defenders (O = offense, X = defense) are at these spots:\n"
        f"{mk_txt}\n\n"
        f"The movements, IN ORDER (move = a player cut/drive, pass = a pass):\n{ar_txt}\n"
        + (f"\nLabels the coach wrote: {', '.join(labels)}\n" if labels else "")
        + "\nReturn STRICT JSON only: {\"name\": \"<short play name, under 40 chars>\", "
        "\"read\": \"<1-2 sentences describing what happens, in order>\"}. No prose."
    )
    import anthropic
    import sys
    client = anthropic.AsyncAnthropic()
    data: dict = {}
    raw = ""
    try:
        resp = await client.messages.create(
            model=OPUS, max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
        blocks = [b for b in resp.content if hasattr(b, "text")]
        raw = (blocks[0].text if blocks else "{}").strip()
        data = _parse_ai_play_json(raw)
    except Exception:
        # One retry: ask for corrected strict JSON.
        try:
            fix = await client.messages.create(
                model=OPUS, max_tokens=2000,
                messages=[{"role": "user", "content":
                    "Return ONLY strict JSON of the form {\"name\":\"...\",\"read\":\"...\"} for this play "
                    "text — no prose, no markdown:\n\n" + (raw or prompt)}],
            )
            fb = [b for b in fix.content if hasattr(b, "text")]
            data = _parse_ai_play_json(fb[0].text if fb else "{}")
        except Exception as exc:
            # Auto-naming is best-effort — never 500; the client just skips naming.
            print(f"ai-play-name: naming skipped ({exc})", file=sys.stderr)
            return {"name": "", "read": ""}
    return {"name": str(data.get("name") or "")[:60], "read": str(data.get("read") or "")[:400]}


@router.delete("/whiteboards/{board_id}")
def delete_whiteboard(
    board_id: int,
    db: Session = Depends(get_db),
    coach: models.Coach = Depends(get_current_coach),
):
    board = db.get(models.GameWhiteboard, board_id)
    if not board or board.coach_id != coach.id:
        raise HTTPException(status_code=404, detail="Whiteboard not found")
    db.delete(board)
    db.commit()
    return {"ok": True}
